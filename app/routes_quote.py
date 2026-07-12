"""Quote route and formula validation."""

import asyncio
import json
import logging
import re
from datetime import datetime
from typing import List, Optional

from fastapi import Depends, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import JSONResponse
from pydantic import BaseModel, Field

from .config import (
    DEFAULT_MATERIALS,
    DEFAULT_PRICING_CONFIG,
    MAX_FILES_PER_REQUEST,
    QUOTE_CONCURRENCY,
    FREE_TOTAL_MODEL_LIMIT,
)
from .db import get_db_session
from .models_orm import User, QuoteHistory
from .deps import get_current_user, get_membership_effective
from .audit import (
    write_audit_event,
    get_idempotency_key_from_request,
    try_get_idempotent_response,
    save_idempotent_response,
)
from .slicer_presets import get_slicer_preset_by_id
from calculator.cost import (
    validate_formula_expression,
    FORMULA_ALIAS_TO_CANONICAL,
    process_single_file,
)

logger = logging.getLogger(__name__)


async def get_quote(
    request: Request,
    files: List[UploadFile] = File(...),
    material: str = Form("PLA", min_length=1, max_length=40),
    layer_height: float = Form(0.2, ge=0.05, le=1.0),
    infill: int = Form(20, ge=0, le=100),
    wall_count: int = Form(3, ge=1, le=20),
    slicer_preset_id: Optional[int] = Form(default=None),
    quantity: int = Form(1, ge=1, le=5000),
    color: str = Form("White", min_length=1, max_length=40),
    use_prusaslicer: Optional[bool] = Form(default=None),
    printer_model: Optional[str] = Form(default=None),
    auto_orient: Optional[bool] = Form(default=False),
    orient_x: Optional[float] = Form(default=None),
    orient_y: Optional[float] = Form(default=None),
    orient_z: Optional[float] = Form(default=None),
    current_user=Depends(get_current_user),
):
    logger.warning("ROUTE_DEBUG auto_orient=%s type=%s", auto_orient, type(auto_orient).__name__)
    from .config import MEMBER_DISCOUNT_PERCENT

    try:
        idem_key = get_idempotency_key_from_request(request)
        if idem_key:
            cached = try_get_idempotent_response(int(current_user["id"]), request, idem_key)
            if cached:
                status_code, payload = cached
                write_audit_event(
                    action="quote.replay",
                    request=request,
                    user=current_user,
                    idempotency_key=idem_key,
                    detail={"status_code": status_code},
                )
                return JSONResponse(status_code=status_code, content=payload)

        if not files:
            raise HTTPException(status_code=400, detail="请至少上传一个模型文件")
        if len(files) > MAX_FILES_PER_REQUEST:
            raise HTTPException(status_code=400, detail=f"单次上传文件数量不能超过 {MAX_FILES_PER_REQUEST} 个")
        # 免费用户累计模型总数限制（不限制单次数量）
        from .deps import is_member_user
        from sqlalchemy import func as sqlfunc

        if not is_member_user(current_user):
            with get_db_session() as db:
                existing_count = (
                    db.query(sqlfunc.count(QuoteHistory.id))
                    .filter(
                        QuoteHistory.user_id == current_user["id"],
                        QuoteHistory.status == "success",
                    )
                    .scalar()
                    or 0
                )
            if existing_count >= FREE_TOTAL_MODEL_LIMIT:
                raise HTTPException(
                    status_code=400, detail=f"免费用户最多累计 {FREE_TOTAL_MODEL_LIMIT} 个模型，升级会员无限制"
                )

        with get_db_session() as db:
            row = db.query(User.materials, User.pricing_config).filter(User.id == current_user["id"]).first()
        user_materials = json.loads(row.materials) if row and row.materials else DEFAULT_MATERIALS
        pricing_config = json.loads(row.pricing_config) if row and row.pricing_config else DEFAULT_PRICING_CONFIG

        if use_prusaslicer is not None:
            pricing_config["use_prusaslicer"] = use_prusaslicer
        if printer_model is not None:
            pricing_config["printer_model"] = printer_model

        material_names = {str(m.get("name")) for m in user_materials if isinstance(m, dict)}
        if material not in material_names:
            raise HTTPException(status_code=400, detail="材料参数不合法")

        selected_material = next(
            (m for m in user_materials if isinstance(m, dict) and str(m.get("name")) == material), None
        )
        allowed_colors = []
        if selected_material:
            raw_colors = selected_material.get("colors", [])
            if isinstance(raw_colors, list):
                for c in raw_colors:
                    if isinstance(c, dict):
                        allowed_colors.append(str(c.get("hex", "")).strip())
                        allowed_colors.append(str(c.get("name", "")).strip())
                    else:
                        allowed_colors.append(str(c).strip())
                allowed_colors = [a for a in allowed_colors if a]
        if allowed_colors and color not in allowed_colors:
            raise HTTPException(status_code=400, detail="颜色参数不合法")

        slicer_preset = None
        if slicer_preset_id is not None:
            sid = int(slicer_preset_id)
            # id=0 (formerly system default) → treated as no preset
            if sid <= 0:
                slicer_preset = None
            else:
                slicer_preset = get_slicer_preset_by_id(int(current_user["id"]), sid)
                if slicer_preset is None:
                    raise HTTPException(status_code=400, detail="切片预设不存在或无权限")

        # Parallel file processing with concurrency limit
        _semaphore = asyncio.Semaphore(max(1, QUOTE_CONCURRENCY))

        async def process_one(file):
            async with _semaphore:
                try:
                    return await asyncio.to_thread(
                        _process_single_file_sync,
                        file,
                        material,
                        layer_height,
                        infill,
                        quantity,
                        color,
                        user_materials,
                        pricing_config,
                        slicer_preset,
                        wall_count,
                        current_user,
                        auto_orient,
                        orient_x,
                        orient_y,
                        orient_z,
                    )
                except Exception as e:
                    fname = file.filename or "unknown"
                    logger.error(f"处理文件 {fname} 时发生未捕获的错误: {str(e)}", exc_info=True)
                    return {
                        "filename": fname,
                        "status": "failed",
                        "error": f"INTERNAL_ERROR: {str(e)}",
                        "cost_cny": 0,
                        "weight_g": 0,
                        "estimated_time_h": 0,
                    }

        tasks = [process_one(f) for f in files]
        results = await asyncio.gather(*tasks)

        success_items = [item for item in results if item.get("status") == "success"]
        failed_items = [item for item in results if item.get("status") == "failed"]

        membership_level, membership_expires_at = get_membership_effective(current_user)
        discount_percent = float(MEMBER_DISCOUNT_PERCENT or 0.0)
        if discount_percent < 0:
            discount_percent = 0.0
        if discount_percent > 90:
            discount_percent = 90.0
        if membership_level == "member" and discount_percent > 0 and success_items:
            for item in success_items:
                try:
                    original = float(item.get("cost_cny") or 0.0)
                except Exception:
                    original = 0.0
                discounted = round(original * (1.0 - (discount_percent / 100.0)), 2)
                item["cost_cny_original"] = round(original, 2)
                item["cost_cny"] = discounted
                breakdown = item.get("cost_breakdown")
                if isinstance(breakdown, dict):
                    breakdown["member_discount_percent"] = round(discount_percent, 2)
                    breakdown["member_discount_cny"] = round(max(0.0, original - discounted), 2)

        payload = {
            "total_files": len(results),
            "success_count": len(success_items),
            "failed_count": len(failed_items),
            "summary_total_cost_cny": round(sum(item.get("cost_cny", 0) for item in success_items), 2),
            "summary_total_weight_g": round(sum(item.get("weight_g", 0) for item in success_items), 2),
            "summary_total_time_h": round(sum(item.get("estimated_time_h", 0) for item in success_items), 2),
            "results": results,
            "membership_level": membership_level,
            "membership_expires_at": membership_expires_at,
            "member_discount_percent": round(discount_percent, 2) if membership_level == "member" else 0.0,
        }
        write_audit_event(
            action="quote.create",
            request=request,
            user=current_user,
            idempotency_key=idem_key,
            detail={
                "files": len(results),
                "success": len(success_items),
                "failed": len(failed_items),
                "material": material,
                "quantity": quantity,
            },
        )
        if idem_key:
            save_idempotent_response(int(current_user["id"]), request, idem_key, 200, payload)
        _save_quote_history(int(current_user["id"]), results)
        return payload
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"处理报价请求失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"INTERNAL_ERROR: 报价请求失败 ({str(e)})")


class FormulaValidateRequest(BaseModel):
    unit_cost_formula: str = Field(..., min_length=1, max_length=800)
    total_cost_formula: str = Field(..., min_length=1, max_length=800)


async def validate_formula(payload: FormulaValidateRequest, current_user=Depends(get_current_user)):
    unit_ok, unit_err, unit_vars = validate_formula_expression(payload.unit_cost_formula)
    total_ok, total_err, total_vars = validate_formula_expression(payload.total_cost_formula)
    ok = unit_ok and total_ok
    return {
        "ok": ok,
        "unit": {"ok": unit_ok, "error": unit_err, "used_vars": unit_vars},
        "total": {"ok": total_ok, "error": total_err, "used_vars": total_vars},
        "aliases": FORMULA_ALIAS_TO_CANONICAL,
    }


# ── Quote history ──


def _process_single_file_sync(
    file: UploadFile,
    material: str,
    layer_height: float,
    infill: int,
    quantity: int,
    color: str,
    user_materials: List[dict],
    pricing_config: dict,
    slicer_preset: Optional[dict] = None,
    perimeters: Optional[int] = None,
    current_user: Optional[dict] = None,
    auto_orient: bool = False,
    orient_x: Optional[float] = None,
    orient_y: Optional[float] = None,
    orient_z: Optional[float] = None,
):
    """Synchronous wrapper for process_single_file — runs in thread pool.

    Uses asyncio.run() which properly creates an event loop, runs the
    coroutine, cancels any remaining async generators/tasks, and closes
    the loop — replacing the old manual loop management that could leak
    resources if the coroutine left pending tasks.
    """
    import asyncio

    return asyncio.run(
        process_single_file(
            file,
            material,
            layer_height,
            infill,
            quantity,
            color,
            user_materials,
            pricing_config,
            slicer_preset,
            perimeters,
            current_user,
            auto_orient,
            orient_x,
            orient_y,
            orient_z,
        )
    )


def _save_quote_history(user_id: int, results: list) -> None:
    """Save quote results to history table."""
    from datetime import datetime, timezone

    now = datetime.now(timezone.utc).isoformat()
    with get_db_session() as db:
        for item in results:
            # Extract new fields
            raw_pm = item.get("_printer_model") or item.get("printer_model") or ""
            slicer_preset_id = item.get("_slicer_preset_id") or item.get("slicer_preset_id")

            breakdown = item.get("cost_breakdown")
            gcode_summary = (breakdown or {}).get("gcode_summary") or {}
            core_params = gcode_summary.get("core_params") or {}

            # nozzle_diameter from slicer params — init before extraction so we can fallback
            nozzle_diameter = None
            if core_params.get("nozzle_diameter") is not None:
                try:
                    nozzle_diameter = float(core_params["nozzle_diameter"])
                except (ValueError, TypeError):
                    pass

            # 从复合ID(bambu_a1_04)提取裸模型ID，喷嘴直径作为fallback
            m = re.match(r"^(.+?)_(\d{2})$", raw_pm) if raw_pm else None
            if m:
                printer_model = m.group(1)  # bambu_a1
                # 查找打印机显示名称
                try:
                    from app.printers import PRINTER_MODELS

                    for pm_def in PRINTER_MODELS:
                        if pm_def["id"] == printer_model:
                            printer_model = pm_def["name"]
                            break
                except Exception:
                    pass  # keep raw ID if resolution fails
                if nozzle_diameter is None:
                    nozzle_diameter = float(m.group(2)) / 10.0  # 04 → 0.4
            else:
                printer_model = raw_pm or None
                # 非复合ID（如user_3），也尝试解析显示名称
                if printer_model:
                    try:
                        from app.printers import PRINTER_MODELS, resolve_printer

                        rp = resolve_printer(printer_model)
                        if rp:
                            printer_model = rp.get("name", printer_model)
                    except Exception:
                        pass

            # layer_height from item directly
            layer_height_val = item.get("layer_height")
            try:
                layer_height_val = float(layer_height_val) if layer_height_val is not None else None
            except (ValueError, TypeError):
                layer_height_val = None

            # wall_count from slicer params perimeters
            wall_count = None
            if core_params.get("perimeters") is not None:
                try:
                    wall_count = int(core_params["perimeters"])
                except (ValueError, TypeError):
                    pass

            # infill from slicer params fill_density (strip %)
            infill_val = None
            raw_fill = core_params.get("fill_density")
            if raw_fill is not None:
                try:
                    infill_val = int(float(str(raw_fill).replace("%", "")))
                except (ValueError, TypeError):
                    pass

            # brand from item
            brand = item.get("brand") or ""

            # cost_breakdown as JSON string
            cost_breakdown_str = json.dumps(breakdown) if isinstance(breakdown, dict) else None

            entry = QuoteHistory(
                user_id=user_id,
                filename=str(item.get("filename") or "")[:200],
                material=str(item.get("material") or "")[:40],
                color=str(item.get("color") or "")[:40],
                quantity=int(item.get("quantity") or 1),
                volume_cm3=round(float(item.get("volume_cm3") or 0), 2),
                weight_g=round(float(item.get("weight_g") or 0), 2),
                estimated_time_h=round(float(item.get("estimated_time_h") or 0), 2),
                cost_cny=round(float(item.get("cost_cny") or 0), 2),
                dimensions=str(item.get("dimensions") or "")[:80],
                status=str(item.get("status") or "success")[:20],
                error_msg=str(item.get("error") or "")[:300] if item.get("status") != "success" else None,
                created_at=now,
                printer_model=str(printer_model)[:50] if printer_model else None,
                slicer_preset_id=int(slicer_preset_id) if slicer_preset_id is not None else None,
                nozzle_diameter=round(float(nozzle_diameter), 2) if nozzle_diameter is not None else None,
                layer_height=round(float(layer_height_val), 2) if layer_height_val is not None else None,
                wall_count=wall_count,
                infill=infill_val,
                brand=str(brand)[:40] if brand else None,
                cost_breakdown=cost_breakdown_str,
                slicer_fallback=int(bool((breakdown or {}).get("slicer_fallback"))) if breakdown else 0,
                slicer_error=(breakdown or {}).get("slicer_error") if breakdown else None,
                slicer_estimated_time_s=float((breakdown or {}).get("slicer_estimated_time_s", 0) or 0)
                if (breakdown or {}).get("slicer_estimated_time_s")
                else None,
            )
            db.add(entry)


async def delete_quote_history(id: int, request: Request, current_user=Depends(get_current_user)):
    """Delete a single quote history record by id. Users can only delete their own records."""
    uid = int(current_user["id"])
    try:
        with get_db_session() as db:
            row = (
                db.query(QuoteHistory)
                .filter(
                    QuoteHistory.id == int(id),
                    QuoteHistory.user_id == uid,
                )
                .first()
            )
            if row is None:
                raise HTTPException(status_code=404, detail="报价记录不存在或无权限删除")
            db.delete(row)
        logger.info(f"用户 {uid} 删除报价记录 id={id}")
        write_audit_event(
            action="quote.history.delete",
            request=request,
            user=current_user,
            detail={"deleted_id": int(id)},
        )
        return {"status": "ok"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除报价记录失败: user_id={uid} id={id} error={str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


async def clear_quote_history(request: Request, current_user=Depends(get_current_user)):
    """Delete all quote history records for the current user."""
    uid = int(current_user["id"])
    try:
        with get_db_session() as db:
            count = db.query(QuoteHistory).filter(QuoteHistory.user_id == uid).delete()
        logger.info(f"用户 {uid} 清理全部报价记录，共删除 {count} 条")
        write_audit_event(
            action="quote.history.clear",
            request=request,
            user=current_user,
            detail={"deleted_count": count},
        )
        return {"status": "ok", "deleted": count}
    except Exception as e:
        logger.error(f"清理报价记录失败: user_id={uid} error={str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"清理失败: {str(e)}")


async def quote_history(limit: int = 20, offset: int = 0, current_user=Depends(get_current_user)):
    """Get quote history for current user."""
    safe_limit = max(1, min(int(limit), 100))
    safe_offset = max(0, int(offset))
    from .deps import is_member_user

    if not is_member_user(current_user):
        safe_limit = min(safe_limit, 10)
    uid = int(current_user["id"])
    with get_db_session() as db:
        total = db.query(QuoteHistory).filter(QuoteHistory.user_id == uid).count()
        rows = (
            db.query(QuoteHistory)
            .filter(QuoteHistory.user_id == uid)
            .order_by(QuoteHistory.id.desc())
            .offset(safe_offset)
            .limit(safe_limit)
            .all()
        )
        items = []
        for r in rows:
            items.append(
                {
                    "id": r.id,
                    "filename": r.filename,
                    "material": r.material,
                    "color": r.color,
                    "quantity": r.quantity,
                    "volume_cm3": round(float(r.volume_cm3 or 0), 2),
                    "weight_g": round(float(r.weight_g or 0), 2),
                    "estimated_time_h": round(float(r.estimated_time_h or 0), 2),
                    "cost_cny": round(float(r.cost_cny or 0), 2),
                    "dimensions": r.dimensions,
                    "status": r.status,
                    "error_msg": r.error_msg,
                    "created_at": r.created_at,
                    "printer_model": r.printer_model,
                    "slicer_preset_id": r.slicer_preset_id,
                    "nozzle_diameter": round(float(r.nozzle_diameter), 2) if r.nozzle_diameter is not None else None,
                    "layer_height": round(float(r.layer_height), 2) if r.layer_height is not None else None,
                    "wall_count": r.wall_count,
                    "infill": r.infill,
                    "brand": r.brand,
                    "cost_breakdown": json.loads(r.cost_breakdown) if r.cost_breakdown else None,
                }
            )
    return {"items": items, "total": total, "limit": safe_limit, "offset": safe_offset}


# ── Quote export ──

_EXPORT_COLUMNS = [
    ("filename", "文件名"),
    ("material", "材料"),
    ("color", "颜色"),
    ("quantity", "数量"),
    ("volume_cm3", "体积(cm³)"),
    ("weight_g", "重量(g)"),
    ("estimated_time_h", "打印时间(h)"),
    ("cost_cny", "成本(元)"),
    ("created_at", "创建时间"),
    ("printer_model", "打印机型号"),
    ("slicer_preset_id", "切片预设ID"),
    ("nozzle_diameter", "喷嘴直径(mm)"),
    ("layer_height", "层高(mm)"),
    ("wall_count", "墙层数"),
    ("infill", "填充(%)"),
    ("brand", "品牌"),
]


def _build_export_query(db, user_id: int, material: Optional[str], date_from: Optional[str], date_to: Optional[str]):
    """Build filtered query for quote history export. Returns a SQLAlchemy Query object."""
    query = db.query(QuoteHistory).filter(QuoteHistory.user_id == user_id)
    if material:
        query = query.filter(QuoteHistory.material == material)
    if date_from:
        # date_from: inclusive, e.g. "2024-01-01" — stored as ISO string
        query = query.filter(QuoteHistory.created_at >= date_from)
    if date_to:
        # date_to: inclusive end-of-day — append "T23:59:59" to cover the full day
        date_to_end = date_to if "T" in date_to else f"{date_to}T23:59:59"
        query = query.filter(QuoteHistory.created_at <= date_to_end)
    return query.order_by(QuoteHistory.id.desc())


def _row_to_export_list(row) -> list:
    """Convert a QuoteHistory ORM row to an export value list."""
    return [
        row.filename or "",
        row.material or "",
        row.color or "",
        int(row.quantity or 0),
        round(float(row.volume_cm3 or 0), 2),
        round(float(row.weight_g or 0), 2),
        round(float(row.estimated_time_h or 0), 2),
        round(float(row.cost_cny or 0), 2),
        str(row.created_at) if row.created_at else "",
        row.printer_model or "",
        int(row.slicer_preset_id) if row.slicer_preset_id is not None else "",
        round(float(row.nozzle_diameter), 2) if row.nozzle_diameter is not None else "",
        round(float(row.layer_height), 2) if row.layer_height is not None else "",
        int(row.wall_count) if row.wall_count is not None else "",
        int(row.infill) if row.infill is not None else "",
        row.brand or "",
    ]


async def export_quote_history(
    request: Request,
    format: str = "csv",
    material: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user=Depends(get_current_user),
):
    """Export quote history as CSV or XLSX file download."""
    import csv
    import io
    from datetime import datetime
    from fastapi.responses import StreamingResponse
    from .deps import is_member_user

    if not is_member_user(current_user):
        raise HTTPException(status_code=403, detail="导出功能仅限会员使用")

    fmt = (format or "csv").strip().lower()
    if fmt not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="format 参数必须为 csv 或 xlsx")

    uid = int(current_user["id"])
    logger.info(f"用户 {uid} 请求导出报价历史 format={fmt} material={material} date_from={date_from} date_to={date_to}")

    try:
        with get_db_session() as db:
            query = _build_export_query(db, uid, material, date_from, date_to)
            rows = query.all()

        if not rows:
            raise HTTPException(status_code=404, detail="没有符合条件的报价记录可导出")

        headers = [col[1] for col in _EXPORT_COLUMNS]
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

        if fmt == "csv":
            buf = io.StringIO()
            # Write UTF-8 BOM for Excel compatibility
            buf.write("\ufeff")
            writer = csv.writer(buf)
            writer.writerow(headers)
            for r in rows:
                writer.writerow(_row_to_export_list(r))
            buf.seek(0)
            filename = f"quote_history_{timestamp}.csv"
            return StreamingResponse(
                buf,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        else:
            # XLSX via openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "报价历史"

            # Header styling
            Font(bold=True, size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(bold=True, size=11, color="FFFFFF")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Write headers
            for col_idx, (_, title) in enumerate(_EXPORT_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=title)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Write data rows
            for row_idx, r in enumerate(rows, 2):
                for col_idx, val in enumerate(_row_to_export_list(r), 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border

            # Auto-fit column widths (approximate)
            for col_idx, (_, title) in enumerate(_EXPORT_COLUMNS, 1):
                max_len = len(title) * 2  # CJK chars are wider
                for row in ws.iter_rows(min_row=2, max_row=min(len(rows) + 1, 102), min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is not None:
                            val_len = len(str(cell.value))
                            max_len = max(max_len, val_len)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            filename = f"quote_history_{timestamp}.xlsx"
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出报价历史失败: user_id={uid} error={str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


# ── PDF Quote Export ──


async def export_quote_pdf(
    request: Request,
    ids: str = "",
    current_user=Depends(get_current_user),
):
    """Export selected quote history items as a branded PDF quote.

    GET /api/quote/export-pdf?ids=1,2,3
    - Requires auth and active membership
    - Accepts comma-separated quote history IDs
    - Validates all IDs belong to current user
    - Generates PDF with brand settings
    """
    from fastapi.responses import Response
    from .deps import is_member_user
    from .pdf_quote import generate_pdf_quote
    from .config import MEMBER_DISCOUNT_PERCENT

    if not is_member_user(current_user):
        raise HTTPException(status_code=403, detail="PDF报价单导出功能仅限会员使用")

    uid = int(current_user["id"])

    # Parse IDs
    if not ids or not ids.strip():
        raise HTTPException(status_code=400, detail="请提供报价记录ID (ids参数)")

    try:
        id_list = [int(x.strip()) for x in ids.split(",") if x.strip()]
    except ValueError:
        raise HTTPException(status_code=400, detail="ids参数格式不正确，应为逗号分隔的数字")

    if not id_list:
        raise HTTPException(status_code=400, detail="请至少提供一个报价记录ID")

    if len(id_list) > 100:
        raise HTTPException(status_code=400, detail="单次最多导出100条记录")

    try:
        with get_db_session() as db:
            # Fetch brand settings
            user_row = db.query(User).filter(User.id == uid).first()
            brand_name = user_row.brand_name if user_row else ""
            brand_logo_url = user_row.brand_logo_url if user_row else ""
            brand_phone = user_row.brand_phone if user_row else ""
            brand_contact_email = user_row.brand_contact_email if user_row else ""
            brand_address = user_row.brand_address if user_row else ""
            brand_note = user_row.brand_note if user_row else ""

            # Fetch quote history items
            rows = (
                db.query(QuoteHistory)
                .filter(
                    QuoteHistory.id.in_(id_list),
                    QuoteHistory.user_id == uid,
                )
                .order_by(QuoteHistory.id.asc())
                .all()
            )

        if not rows:
            raise HTTPException(status_code=404, detail="未找到指定的报价记录")

        # Build items list
        items = []
        for r in rows:
            items.append(
                {
                    "filename": r.filename or "",
                    "material": r.material or "",
                    "color": r.color or "",
                    "quantity": int(r.quantity or 1),
                    "volume_cm3": round(float(r.volume_cm3 or 0), 2),
                    "weight_g": round(float(r.weight_g or 0), 2),
                    "estimated_time_h": round(float(r.estimated_time_h or 0), 2),
                    "cost_cny": round(float(r.cost_cny or 0), 2),
                    "printer_model": r.printer_model or "",
                    "nozzle_diameter": round(float(r.nozzle_diameter), 2) if r.nozzle_diameter is not None else "",
                    "layer_height": round(float(r.layer_height), 2) if r.layer_height is not None else 0,
                    "wall_count": int(r.wall_count) if r.wall_count is not None else 0,
                    "infill_percent": int(r.infill) if r.infill is not None else 0,
                    "brand": r.brand or "",
                }
            )

        # Determine member discount
        membership_level, _ = get_membership_effective(current_user)
        discount_pct = 0.0
        if membership_level == "member":
            discount_pct = float(MEMBER_DISCOUNT_PERCENT or 0.0)
            discount_pct = max(0.0, min(90.0, discount_pct))

        # Generate PDF
        pdf_bytes = generate_pdf_quote(
            items=items,
            brand_name=brand_name or "",
            brand_logo_url=brand_logo_url or "",
            brand_phone=brand_phone or "",
            brand_contact_email=brand_contact_email or "",
            brand_address=brand_address or "",
            brand_note=brand_note or "",
            member_discount_percent=discount_pct,
        )

        # Build filename
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"quote_{timestamp}.pdf"

        logger.info(f"用户 {uid} 导出PDF报价单，包含 {len(items)} 个项目")

        write_audit_event(
            action="quote.export_pdf",
            request=request,
            user=current_user,
            detail={"ids": id_list, "count": len(items)},
        )

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(pdf_bytes)),
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出PDF报价单失败: user_id={uid} error={str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"PDF导出失败: {str(e)}")


class PdfInlineItem(BaseModel):
    filename: str = ""
    material: str = ""
    color: str = ""
    quantity: int = 1
    volume_cm3: float = 0
    weight_g: float = 0
    estimated_time_h: float = 0
    cost_cny: float = 0
    created_at: str = ""
    status: str = "success"
    # Additional model parameters
    printer_model: str = ""
    nozzle_diameter: str = ""
    layer_height: float = 0
    wall_count: int = 0
    infill_percent: int = 0
    brand: str = ""
    thumbnail_b64: str = ""


class PdfInlineRequest(BaseModel):
    items: List[PdfInlineItem]


async def export_pdf_inline(
    payload: PdfInlineRequest,
    request: Request,
    current_user=Depends(get_current_user),
):
    """POST /api/quote/export-pdf-inline — 从当前页面结果直接生成PDF（不查DB）"""
    from fastapi.responses import Response
    from .pdf_quote import generate_pdf_quote
    from .deps import is_member_user, get_membership_effective
    from .config import MEMBER_DISCOUNT_PERCENT

    if not is_member_user(current_user):
        raise HTTPException(status_code=403, detail="会员专属功能")

    if not payload.items:
        raise HTTPException(status_code=400, detail="没有可导出的项目")

    # Fetch brand settings
    uid = int(current_user["id"])
    brand_name = brand_logo_url = brand_phone = brand_contact_email = brand_address = brand_note = ""
    with get_db_session() as db:
        user = db.query(User).filter(User.id == uid).first()
        if user:
            brand_name = user.brand_name or ""
            brand_logo_url = user.brand_logo_url or ""
            brand_phone = user.brand_phone or ""
            brand_contact_email = user.brand_contact_email or ""
            brand_address = user.brand_address or ""
            brand_note = user.brand_note or ""

    level, _ = get_membership_effective(current_user)
    discount_pct = float(MEMBER_DISCOUNT_PERCENT or 0.0) if level == "member" else 0.0
    discount_pct = max(0.0, min(90.0, discount_pct))

    items = [item.model_dump() for item in payload.items]
    pdf_bytes = generate_pdf_quote(
        items=items,
        brand_name=brand_name,
        brand_logo_url=brand_logo_url,
        brand_phone=brand_phone,
        brand_contact_email=brand_contact_email,
        brand_address=brand_address,
        brand_note=brand_note,
        member_discount_percent=discount_pct,
    )

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"quote_{timestamp}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
        },
    )
