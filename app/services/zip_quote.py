"""ZIP quote business services."""

import asyncio
import io
import json
import logging
import os
import re
import time
import uuid
import zipfile
from typing import Optional

from fastapi import HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.audit import write_audit_event
from app.config import (
    DEFAULT_COLORS,
    DEFAULT_MATERIALS,
    FREE_TOTAL_MODEL_LIMIT,
    MAX_FILES_PER_REQUEST,
    MAX_ZIP_SIZE_BYTES,
    SUPPORTED_EXTENSIONS,
)
from app.db import get_db_session
from app.deps import is_member_user
from app.models_orm import QuoteHistory, User
from app.zip_parser import (
    _collect_all_warnings,
    _match_checklist_to_models,
    _parse_excel_checklist,
)
from app.services.quote import (
    _load_user_quote_settings,
    _process_single_file_sync,
    _resolve_effective_printer_model,
    _resolve_effective_slicer_params,
    _resolve_effective_slicer_preset,
    save_quote_history,
)
from app.material_resolver import merge_user_material_with_catalog

logger = logging.getLogger(__name__)


_COLOR_NAME_TO_HEX = {}
try:
    for _c in DEFAULT_COLORS:
        if isinstance(_c, dict) and _c.get("name") and _c.get("hex"):
            _COLOR_NAME_TO_HEX[_c["name"].strip()] = _c["hex"].strip()
except Exception:
    pass

_ENGLISH_COLOR_MAP = {
    "White": "#ffffff",
    "Black": "#000000",
    "Gray": "#808080",
    "Grey": "#808080",
    "Red": "#dc2626",
    "Blue": "#2563eb",
    "Green": "#16a34a",
    "Yellow": "#ca8a04",
    "Orange": "#ea580c",
    "Purple": "#9333ea",
    "Pink": "#db2777",
}
_COLOR_NAME_TO_HEX.update(_ENGLISH_COLOR_MAP)


_preview_sessions = {}
_PREVIEW_SESSION_TTL = 600
_UPLOAD_READ_CHUNK_BYTES = 1024 * 1024


def _match_selected_material(user_materials: list, material_name: str, brand: str = "", color: str = "") -> Optional[dict]:
    material_name = str(material_name or "").strip()
    brand = str(brand or "").strip()
    color = str(color or "").strip().lower()
    if not material_name:
        return None

    candidates = [
        m
        for m in user_materials
        if isinstance(m, dict)
        and str(m.get("name") or "").strip() == material_name
        and (not brand or str(m.get("brand") or "Generic").strip() == brand)
    ]
    if not candidates:
        return merge_user_material_with_catalog(None, material_name, brand, color)
    if color:
        for candidate in candidates:
            raw_color = candidate.get("color")
            values = []
            if isinstance(raw_color, dict):
                values.extend([raw_color.get("hex"), raw_color.get("name")])
            elif raw_color:
                values.append(raw_color)
            if any(str(v or "").strip().lower() == color for v in values):
                return merge_user_material_with_catalog(candidate, material_name, brand, color)
    return merge_user_material_with_catalog(candidates[0], material_name, brand, color)


def _resolve_color_hex(color_str: str, fallback: str = "") -> str:
    """Convert a color name or hex to a valid hex string."""
    s = color_str.strip()
    if not s:
        return fallback
    if s.startswith("#") and len(s) == 7:
        return s
    hex_val = _COLOR_NAME_TO_HEX.get(s)
    if hex_val:
        return hex_val
    for name, hx in _COLOR_NAME_TO_HEX.items():
        if name in s or s in name:
            return hx
    return fallback or s


def _color_tokens(value) -> set[str]:
    if isinstance(value, dict):
        values = (value.get("name"), value.get("hex"))
    else:
        values = (value,)
    return {str(item or "").strip().lower() for item in values if str(item or "").strip()}


def _build_missing_checklist_materials(user_materials: list, checklist: Optional[list]) -> list:
    """Build Generic material-color records required by explicit checklist colors."""
    if not checklist:
        return []

    materials = [item for item in user_materials if isinstance(item, dict)]
    created = []
    for item in checklist:
        color_name = str(item.get("color") or "").strip()
        if not color_name:
            continue
        material_name = str(item.get("material_type") or "").strip() or "PLA"
        resolved_color = _resolve_color_hex(color_name)
        color_hex = resolved_color if re.fullmatch(r"#[0-9a-fA-F]{6}", resolved_color or "") else ""
        wanted_tokens = _color_tokens({"name": color_name, "hex": color_hex})

        exists = any(
            str(material.get("brand") or "Generic").strip().lower() == "generic"
            and str(material.get("name") or "").strip().lower() == material_name.lower()
            and bool(_color_tokens(material.get("color")) & wanted_tokens)
            for material in [*materials, *created]
        )
        if exists:
            continue

        base = next(
            (
                material
                for material in materials
                if str(material.get("brand") or "Generic").strip().lower() == "generic"
                and str(material.get("name") or "").strip().lower() == material_name.lower()
            ),
            None,
        )
        if base is None:
            base = next(
                (
                    material
                    for material in materials
                    if str(material.get("name") or "").strip().lower() == material_name.lower()
                ),
                None,
            )
        if base is None:
            base = next(
                (
                    material
                    for material in DEFAULT_MATERIALS
                    if str(material.get("name") or "").strip().lower() == material_name.lower()
                ),
                None,
            )
        if base is None:
            base = next(
                (material for material in DEFAULT_MATERIALS if str(material.get("name") or "").upper() == "PLA"),
                {},
            )

        new_material = {
            key: value
            for key, value in dict(base).items()
            if key not in {"name", "brand", "color"}
        }
        new_material.update(
            {
                "name": material_name,
                "brand": "Generic",
                "color": {"name": color_name, "hex": color_hex},
            }
        )
        created.append(new_material)
    return created


def _ensure_checklist_material_colors(user_id: int, user_materials: list, checklist: Optional[list]) -> list:
    created = _build_missing_checklist_materials(user_materials, checklist)
    if not created:
        return []

    user_materials.extend(created)
    with get_db_session() as db:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="用户不存在")
        user.materials = json.dumps(user_materials, ensure_ascii=False)
    logger.info(
        "ZIP material library updated: user_id=%s created=%s",
        user_id,
        [
            {
                "brand": material.get("brand"),
                "name": material.get("name"),
                "color": material.get("color"),
            }
            for material in created
        ],
    )
    return created


def _store_preview_session(data: dict) -> str:
    """Store parsed ZIP data in memory and return a session_id."""
    session_id = uuid.uuid4().hex
    _preview_sessions[session_id] = {
        "data": data,
        "expires_at": time.time() + _PREVIEW_SESSION_TTL,
    }
    now = time.time()
    expired = [k for k, v in _preview_sessions.items() if v["expires_at"] < now]
    for k in expired:
        _preview_sessions.pop(k, None)
    return session_id


def _get_preview_session(session_id: str) -> Optional[dict]:
    """Retrieve stored preview data by session_id. Returns None if expired/missing."""
    entry = _preview_sessions.get(session_id)
    if not entry:
        return None
    if entry["expires_at"] < time.time():
        _preview_sessions.pop(session_id, None)
        return None
    return entry["data"]


def _consume_preview_session(session_id: str) -> Optional[dict]:
    """Retrieve and delete stored preview data (one-time use)."""
    data = _get_preview_session(session_id)
    if data:
        _preview_sessions.pop(session_id, None)
    return data


async def _read_upload_limited(file: UploadFile) -> bytes:
    """Read an upload without allocating beyond the configured ZIP limit."""
    chunks = []
    total = 0
    while True:
        chunk = await file.read(_UPLOAD_READ_CHUNK_BYTES)
        if not chunk:
            break
        total += len(chunk)
        if total > MAX_ZIP_SIZE_BYTES:
            raise HTTPException(
                status_code=400,
                detail=f"ZIP 文件不能超过 {MAX_ZIP_SIZE_BYTES // (1024 * 1024)}MB",
            )
        chunks.append(chunk)
    return b"".join(chunks)


def _read_upload_file_limited(file: UploadFile) -> bytes:
    """Read the underlying spooled file for the synchronous fallback path."""
    if not getattr(file, "file", None):
        raise HTTPException(status_code=400, detail="ZIP 文件无法读取")
    try:
        file.file.seek(0)
        content = file.file.read(MAX_ZIP_SIZE_BYTES + 1)
        file.file.seek(0)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"ZIP 文件无法读取：{e}") from e
    if len(content) > MAX_ZIP_SIZE_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"ZIP 文件不能超过 {MAX_ZIP_SIZE_BYTES // (1024 * 1024)}MB",
        )
    return content


def _validate_free_zip_capacity(existing_count: int, incoming_count: int) -> None:
    if existing_count + incoming_count <= FREE_TOTAL_MODEL_LIMIT:
        return
    remaining = max(0, FREE_TOTAL_MODEL_LIMIT - existing_count)
    raise HTTPException(
        status_code=400,
        detail=(
            f"免费用户最多累计 {FREE_TOTAL_MODEL_LIMIT} 个模型，"
            f"当前还可上传 {remaining} 个，本次 ZIP 包含 {incoming_count} 个"
        ),
    )


def _parse_zip_contents(file_bytes: bytes) -> dict:
    """Parse ZIP bytes and return {excel_bytes, stl_files, checklist, match_result}."""
    try:
        zf = zipfile.ZipFile(io.BytesIO(file_bytes))
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="无效的 ZIP 文件，请检查文件完整性")

    entries = [entry for entry in zf.infolist() if not entry.is_dir()]
    expanded_size = sum(entry.file_size for entry in entries)
    if expanded_size > MAX_ZIP_SIZE_BYTES:
        zf.close()
        raise HTTPException(
            status_code=400,
            detail=f"ZIP 解压后总大小不能超过 {MAX_ZIP_SIZE_BYTES // (1024 * 1024)}MB",
        )
    for entry in entries:
        if entry.flag_bits & 0x1:
            zf.close()
            raise HTTPException(status_code=400, detail="ZIP 中不能包含加密文件")
        if entry.file_size > MAX_ZIP_SIZE_BYTES:
            zf.close()
            raise HTTPException(
                status_code=400,
                detail=f"ZIP 中单个文件解压后不能超过 {MAX_ZIP_SIZE_BYTES // (1024 * 1024)}MB",
            )
        if entry.file_size and entry.compress_size == 0:
            zf.close()
            raise HTTPException(status_code=400, detail="ZIP 中包含异常压缩条目")
    excel_bytes = None
    excel_filename = None
    stl_files = []
    model_keys = set()

    for entry in entries:
        bn = os.path.basename(entry.filename)
        path_parts = entry.filename.replace("\\", "/").split("/")
        if bn.startswith(".") or "__MACOSX" in path_parts or bn.startswith("~$"):
            continue

        entry_bytes = zf.read(entry)
        lower = bn.lower()

        if lower.endswith((".xlsx", ".xls")):
            if excel_bytes is not None:
                zf.close()
                raise HTTPException(status_code=400, detail="ZIP 中只能包含一个 Excel 清单")
            excel_bytes = entry_bytes
            excel_filename = bn
            logger.info("ZIP: found Excel checklist: %s", bn)
        else:
            ext = os.path.splitext(lower)[1]
            if ext in SUPPORTED_EXTENSIONS:
                stem = os.path.splitext(bn)[0].lower()
                if stem in model_keys:
                    zf.close()
                    raise HTTPException(
                        status_code=400,
                        detail=f"ZIP 中存在重名模型：{bn}。模型文件名（不含扩展名）必须唯一",
                    )
                model_keys.add(stem)
                stl_files.append(
                    {
                        "filename": bn,
                        "name_stem": stem,
                        "file_bytes": entry_bytes,
                        "ext": ext,
                    }
                )
                logger.info(f"ZIP: found model file: {bn} (stem={stem})")

    zf.close()

    if not stl_files:
        raise HTTPException(status_code=400, detail="ZIP 中未找到支持的模型文件（.stl/.stp/.step/.obj/.3mf）")
    if len(stl_files) > MAX_FILES_PER_REQUEST:
        raise HTTPException(status_code=400, detail=f"ZIP 中模型文件数量不能超过 {MAX_FILES_PER_REQUEST} 个")

    try:
        checklist = _parse_excel_checklist(excel_bytes, excel_filename) if excel_bytes else None
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    if checklist:
        match_result = _match_checklist_to_models(checklist, stl_files)
    else:
        match_result = {
            "matched": [],
            "checklist_only": [],
            "stl_only": stl_files,
            "match_mode": "none",
        }

    return {
        "excel_bytes": excel_bytes,
        "excel_filename": excel_filename,
        "stl_files": stl_files,
        "checklist": checklist,
        "match_result": match_result,
    }


def _resolved_nozzle(printer_id: Optional[str]) -> Optional[float]:
    """Return the nozzle encoded by a compound printer id."""
    if not printer_id:
        return None
    from app.printers import resolve_printer

    resolved = resolve_printer(str(printer_id))
    if not resolved or resolved.get("_nozzle") is None:
        return None
    return float(resolved["_nozzle"])


def _lookup_printer(printer_name, nozzle_str, fallback_nozzle=None):
    """Return compound_id (e.g. bambu_a1_04) or None if not found."""
    from app.printers import PRINTER_MODELS, resolve_printer

    if not printer_name or not str(printer_name).strip():
        return None
    name_lower = str(printer_name).strip().lower()
    for pm in PRINTER_MODELS:
        if pm.get("name", "").lower() == name_lower:
            pid = pm["id"]
            nz = fallback_nozzle
            if nozzle_str and str(nozzle_str).strip():
                try:
                    nz = float(str(nozzle_str).strip())
                except (ValueError, TypeError):
                    pass
            resolved = resolve_printer(pid, nz)
            if resolved:
                return resolved.get("_compound_id") or pid
            return pid
    return None


def _resolve_checklist_printer(
    default_compound_id: Optional[str],
    printer_name: str = "",
    nozzle_str: str = "",
) -> Optional[str]:
    """Apply only checklist printer fields that are explicitly populated."""
    from app.printers import resolve_printer

    printer_name = str(printer_name or "").strip()
    nozzle_str = str(nozzle_str or "").strip()
    default_nozzle = _resolved_nozzle(default_compound_id)

    if printer_name:
        return _lookup_printer(printer_name, nozzle_str, default_nozzle) or default_compound_id
    if not nozzle_str or not default_compound_id:
        return default_compound_id

    try:
        requested_nozzle = float(nozzle_str)
    except (TypeError, ValueError):
        return default_compound_id
    default_printer = resolve_printer(default_compound_id)
    if not default_printer:
        return default_compound_id
    resolved = resolve_printer(default_printer["id"], requested_nozzle)
    return (resolved.get("_compound_id") if resolved else None) or default_compound_id


def _zip_preview_model_path(result: dict, model: dict) -> Optional[str]:
    """Return the ZIP-owned model path exposed by the authenticated download route."""
    return model.get("_pre_saved_path") or result.get("_saved_path")


async def build_zip_preview_response(file: UploadFile):
    """Parse ZIP and return match analysis without slicing."""
    fname = (file.filename or "").lower()
    if not fname.endswith(".zip"):
        raise HTTPException(status_code=400, detail="请上传 .zip 压缩文件")

    content = await _read_upload_limited(file)

    parsed = _parse_zip_contents(content)
    match_result = parsed["match_result"]
    parsed["stl_files"]
    parsed["checklist"]

    matched_list = []
    for m in match_result["matched"]:
        matched_list.append(
            {
                "filename": m["stl"]["filename"],
                "checklist": m["checklist"],
            }
        )

    bom_only_list = []
    for c in match_result["checklist_only"]:
        bom_only_list.append(
            {
                "filename": c.get("filename", c.get("filename_stem", "")),
                "reason": "清单中有但无对应模型",
            }
        )

    model_only_list = []
    for s in match_result["stl_only"]:
        model_only_list.append(
            {
                "filename": s["filename"],
                "reason": "模型不在清单中，将使用默认参数",
            }
        )

    session_id = _store_preview_session(
        {
            "file_bytes": content,
            "stl_files": parsed["stl_files"],
            "checklist": parsed["checklist"],
            "match_result": parsed["match_result"],
            "excel_bytes": parsed["excel_bytes"],
        }
    )

    return JSONResponse(
        {
            "matched": matched_list,
            "bom_only": bom_only_list,
            "model_only": model_only_list,
            "match_summary": {
                "matched": len(matched_list),
                "bom_only": len(bom_only_list),
                "model_only": len(model_only_list),
            },
            "session_id": session_id,
        }
    )


def _resolve_zip_defaults(current_user: dict, file: Optional[UploadFile], session_id: Optional[str]):
    """Resolve preview session or parse uploaded file, plus user defaults."""

    if session_id:
        preview_data = _consume_preview_session(session_id)
        if not preview_data:
            raise HTTPException(status_code=400, detail="预览会话已过期或不存在，请重新上传")
        stl_files = preview_data["stl_files"]
        checklist = preview_data["checklist"]
        match_result = preview_data["match_result"]
        content = preview_data["file_bytes"]
    else:
        if not file:
            raise HTTPException(status_code=400, detail="请上传 .zip 压缩文件或提供 session_id")
        fname = (file.filename or "").lower()
        if not fname.endswith(".zip"):
            raise HTTPException(status_code=400, detail="请上传 .zip 压缩文件")

        content = _read_upload_file_limited(file)
        parsed = _parse_zip_contents(content)
        stl_files = parsed["stl_files"]
        checklist = parsed["checklist"]
        match_result = parsed["match_result"]

    (
        user_materials,
        pricing_config,
        default_printer_id,
        default_nozzle,
        default_slicer_preset_id,
    ) = _load_user_quote_settings(int(current_user["id"]))

    return (
        stl_files,
        checklist,
        match_result,
        content,
        user_materials,
        pricing_config,
        default_printer_id,
        default_nozzle,
        default_slicer_preset_id,
    )


async def build_zip_quote_response(
    request: Request,
    file: Optional[UploadFile],
    material: str,
    color: str,
    quantity: int,
    printer_model: Optional[str],
    slicer_preset_id: Optional[int],
    layer_height: float,
    wall_count: int,
    infill: int,
    session_id: Optional[str],
    current_user: dict,
):
    """Generate ZIP quote streaming response."""
    from app.utils import _user_base_dir
    from sqlalchemy import func as sqlfunc

    (
        stl_files,
        checklist,
        match_result,
        content,
        user_materials,
        pricing_config,
        default_printer_id,
        default_nozzle,
        default_slicer_preset_id,
    ) = _resolve_zip_defaults(current_user, file, session_id)
    user_materials = [dict(material) for material in user_materials if isinstance(material, dict)]

    if not is_member_user(current_user):
        with get_db_session() as db:
            existing_count = (
                db.query(sqlfunc.count(QuoteHistory.id))
                .filter(
                    QuoteHistory.user_id == current_user["id"],
                    QuoteHistory.status == "success",
                )
                .scalar()
                or 0
            )
        _validate_free_zip_capacity(existing_count, len(stl_files))

    created_materials = _ensure_checklist_material_colors(
        int(current_user["id"]),
        user_materials,
        checklist,
    )

    # Pre-save all model files to disk so thumbnails work even if slicing fails
    _user_folder = f"user_{current_user['id']}_{current_user['username']}"
    _zip_job_id = uuid.uuid4().hex[:8]
    _zip_uploads_dir = os.path.join(_user_base_dir(), _user_folder, "uploads", _zip_job_id)
    os.makedirs(_zip_uploads_dir, exist_ok=True)
    for _sf in stl_files:
        _saved = os.path.join(_zip_uploads_dir, _sf["filename"])
        with open(_saved, "wb") as _f:
            _f.write(_sf["file_bytes"])
        _sf["_pre_saved_path"] = _saved

    total_stl = len(stl_files)
    if match_result["match_mode"] == "all":
        match_msg = f"全部模型预设生效（{len(match_result['matched'])}/{total_stl} 个文件匹配）"
    elif match_result["match_mode"] == "partial":
        match_msg = f"部分模型预设生效，请检查清单（{len(match_result['matched'])} 匹配 / {len(match_result['checklist_only'])} 清单多余 / {len(match_result['stl_only'])} 无预设）"
    else:
        if checklist:
            match_msg = f"全部模型预设未生效，请检查清单（{len(stl_files)} 个模型均未匹配）"
        else:
            match_msg = "未包含 Excel 清单，使用默认参数"

    # Resolve defaults with the exact same precedence as direct uploads.
    _default_compound_id = _resolve_effective_printer_model(printer_model, default_printer_id, default_nozzle)
    _, _default_preset = _resolve_effective_slicer_preset(
        int(current_user["id"]),
        slicer_preset_id,
        default_slicer_preset_id,
    )

    # Keep ZIP uploads on the same parameter contract as normal uploads. A
    # preset's core values win over stale form fallbacks, while checklist
    # fields below can still override them per model.
    (
        effective_layer_height,
        effective_wall_count,
        effective_infill,
    ) = _resolve_effective_slicer_params(layer_height, wall_count, infill, _default_preset)

    _files_to_process = []
    for m in match_result["matched"]:
        _files_to_process.append(("matched", m))
    for stl in match_result["stl_only"]:
        _files_to_process.append(("stl_only", stl))
    _total_files = len(_files_to_process)

    async def _generate():
        results = []
        for _idx, (_file_type, _item) in enumerate(_files_to_process, 1):
            if await request.is_disconnected():
                logger.info("ZIP processing cancelled by client")
                yield f"data: {json.dumps({'type': 'cancelled', 'processed': _idx})}\n\n"
                return
            try:
                if _file_type == "matched":
                    cl = _item["checklist"]
                    stl = _item["stl"]

                    _cl_lh_raw = cl.get("layer_height_parsed")
                    _cl_wc_raw = cl.get("wall_count_parsed")
                    _cl_inf_raw = cl.get("infill_parsed")
                    has_print_params = any([_cl_lh_raw, _cl_wc_raw, _cl_inf_raw])

                    if _cl_lh_raw:
                        try:
                            lh = float(_cl_lh_raw)
                        except (ValueError, TypeError):
                            lh = None
                    else:
                        lh = None
                    if _cl_wc_raw:
                        try:
                            wc = int(float(_cl_wc_raw))
                        except (ValueError, TypeError):
                            wc = None
                    else:
                        wc = None
                    if _cl_inf_raw:
                        try:
                            inf = int(float(_cl_inf_raw))
                        except (ValueError, TypeError):
                            inf = None
                    else:
                        inf = None

                    cl_qty = cl.get("quantity_parsed", quantity)
                    cl_color_raw = cl.get("color", "").strip()
                    cl_color = _resolve_color_hex(cl_color_raw, _resolve_color_hex(color))
                    cl_material = cl.get("material_type", "").strip() or material
                    cl_brand = cl.get("material_brand", "").strip()
                    cl_material_spec = _match_selected_material(user_materials, cl_material, cl_brand, cl_color_raw or cl_color)

                    cl_printer = str(cl.get("printer_model", "")).strip()
                    cl_nozzle = str(cl.get("nozzle", "")).strip()
                    has_printer_override = bool(cl_printer)
                    has_nozzle_override = bool(cl_nozzle)
                    compound_id = _resolve_checklist_printer(_default_compound_id, cl_printer, cl_nozzle)

                    file_pricing = dict(pricing_config)
                    if compound_id:
                        file_pricing["printer_model"] = compound_id

                    if has_print_params:
                        file_preset = None
                        _lh = lh if lh is not None else effective_layer_height
                        _wc = wc if wc is not None else effective_wall_count
                        _inf = inf if inf is not None else effective_infill
                    else:
                        file_preset = _default_preset
                        _lh = effective_layer_height
                        _wc = effective_wall_count
                        _inf = effective_infill

                    fake_file = UploadFile(
                        filename=stl["filename"],
                        file=io.BytesIO(stl["file_bytes"]),
                    )

                    result = await asyncio.to_thread(
                        _process_single_file_sync,
                        fake_file,
                        material=cl_material,
                        layer_height=_lh,
                        infill=_inf,
                        quantity=cl_qty,
                        color=cl_color,
                        user_materials=user_materials,
                        pricing_config=file_pricing,
                        slicer_preset=file_preset,
                        perimeters=_wc,
                        current_user=current_user,
                        auto_orient=False,
                        selected_material_spec=cl_material_spec,
                    )
                    result["_checklist_params"] = True
                    result["brand"] = (
                        str(cl_material_spec.get("brand") or cl_brand or "Generic")
                        if cl_material_spec
                        else (cl_brand or "Generic")
                    )
                    result["_printer_model_explicit"] = has_printer_override or has_nozzle_override
                    result["_slicer_preset_explicit"] = has_print_params
                    result["_checklist_source"] = {
                        "layer_height": _lh if has_print_params else "",
                        "wall_count": _wc if has_print_params else "",
                        "infill": _inf if has_print_params else "",
                        "printer_model": cl_printer,
                        "nozzle": cl_nozzle,
                        "material_type": cl.get("material_type", ""),
                        "material_brand": cl.get("material_brand", ""),
                        "color": cl_color_raw,
                        "quantity": cl_qty,
                    }
                    result["checklist_file_path"] = _zip_preview_model_path(result, stl)

                    filename = stl["filename"]
                    pre_saved = stl.get("_pre_saved_path")

                else:
                    stl = _item
                    fake_file = UploadFile(
                        filename=stl["filename"],
                        file=io.BytesIO(stl["file_bytes"]),
                    )
                    file_pricing = dict(pricing_config)
                    if _default_compound_id:
                        file_pricing["printer_model"] = _default_compound_id

                    result = await asyncio.to_thread(
                        _process_single_file_sync,
                        fake_file,
                        material=material,
                        layer_height=effective_layer_height,
                        infill=effective_infill,
                        quantity=quantity,
                        color=_resolve_color_hex(color),
                        user_materials=user_materials,
                        pricing_config=file_pricing,
                        slicer_preset=_default_preset,
                        perimeters=effective_wall_count,
                        current_user=current_user,
                        auto_orient=False,
                        selected_material_spec=_match_selected_material(user_materials, material, "", color),
                    )
                    result["_checklist_params"] = False
                    result["checklist_file_path"] = _zip_preview_model_path(result, stl)

                    filename = stl["filename"]
                    pre_saved = stl.get("_pre_saved_path")

                if not result.get("checklist_file_path") and pre_saved:
                    result["checklist_file_path"] = pre_saved

                results.append(result)

                status = "success" if result.get("status") == "success" else "failed"
                yield f"data: {json.dumps({'type': 'progress', 'current': _idx, 'total': _total_files, 'filename': filename, 'status': status})}\n\n"

            except Exception as e:
                filename = "unknown"
                pre_saved = None
                if isinstance(_item, dict):
                    if _file_type == "matched":
                        filename = _item.get("stl", {}).get("filename", "unknown")
                        pre_saved = _item.get("stl", {}).get("_pre_saved_path")
                    else:
                        filename = _item.get("filename", "unknown")
                        pre_saved = _item.get("_pre_saved_path")
                result = {
                    "filename": filename,
                    "status": "failed",
                    "error": str(e),
                    "cost_cny": 0,
                    "weight_g": 0,
                    "estimated_time_h": 0,
                }
                if pre_saved:
                    result["checklist_file_path"] = pre_saved
                results.append(result)
                yield f"data: {json.dumps({'type': 'progress', 'current': _idx, 'total': _total_files, 'filename': filename, 'status': 'failed'})}\n\n"

        success_items = [r for r in results if r.get("status") == "success"]
        failed_items = [r for r in results if r.get("status") == "failed"]

        payload = {
            "total_files": len(results),
            "success_count": len(success_items),
            "failed_count": len(failed_items),
            "summary_total_cost_cny": round(sum(r.get("cost_cny", 0) for r in success_items), 2),
            "summary_total_weight_g": round(sum(r.get("weight_g", 0) for r in success_items), 2),
            "summary_total_time_h": round(sum(r.get("estimated_time_h", 0) for r in success_items), 2),
            "results": results,
            "created_materials": created_materials,
            "match_status": {
                "mode": match_result["match_mode"],
                "message": match_msg,
                "matched_count": len(match_result["matched"]),
                "checklist_only_count": len(match_result["checklist_only"]),
                "stl_only_count": len(match_result["stl_only"]),
                "checklist_only_files": [
                    c.get("filename", c.get("filename_stem", "")) for c in match_result["checklist_only"]
                ],
                "warnings": _collect_all_warnings(checklist),
            },
        }

        save_quote_history(int(current_user["id"]), results)
        write_audit_event(
            action="quote.zip_upload",
            request=request,
            user=current_user,
            detail={
                "files": len(results),
                "success": len(success_items),
                "failed": len(failed_items),
                "match_mode": match_result["match_mode"],
                "material": material,
                "quantity": quantity,
            },
        )

        yield f"data: {json.dumps({'type': 'done', **payload})}\n\n"

    return StreamingResponse(_generate(), media_type="text/event-stream")


def download_zip_model(file_path: str, current_user: dict):
    """Download a model file that was saved during ZIP processing."""
    from app.utils import _user_base_dir

    user_folder = f"user_{current_user['id']}_{current_user['username']}"
    allowed_dir = os.path.realpath(os.path.join(_user_base_dir(), user_folder, "uploads"))
    abs_path = os.path.realpath(file_path)
    try:
        is_allowed = os.path.commonpath([abs_path, allowed_dir]) == allowed_dir
    except ValueError:
        is_allowed = False
    if not is_allowed:
        raise HTTPException(status_code=403, detail="Access denied")
    if not os.path.isfile(abs_path):
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(abs_path, filename=os.path.basename(abs_path))


def download_zip_template(request: Request):
    """Generate and return an xlsx template for ZIP import checklist."""
    from jose import jwt as jose_jwt
    from app.config import JWT_SECRET_KEY, JWT_ALGORITHM

    user_brands = None
    authorization = request.headers.get("authorization")
    if authorization:
        try:
            token = authorization.replace("Bearer ", "")
            payload = jose_jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
            user_id = int(payload.get("sub", "0"))
            if user_id > 0:
                with get_db_session() as db:
                    u = db.query(User.materials).filter(User.id == user_id).first()
                if u and u.materials:
                    materials = json.loads(u.materials)
                    brands = sorted(
                        {m.get("brand", "Generic") for m in materials if isinstance(m, dict) and m.get("brand")}
                    )
                    if brands:
                        user_brands = brands
        except Exception:
            pass

    if not user_brands:
        user_brands = ["eSUN", "Generic", "Hatchbox", "Polymaker", "Sunlu"]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    sub_header_font = Font(name="Microsoft YaHei", bold=True, size=11)
    sub_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    normal_font = Font(name="Microsoft YaHei", size=11)
    note_font = Font(name="Microsoft YaHei", size=10, color="666666")
    note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "导入模板"

    headers_en = [
        "filename",
        "material_brand",
        "material_type",
        "color",
        "quantity",
        "printer",
        "nozzle",
        "layer_height",
        "wall_count",
        "infill",
    ]
    for col, val in enumerate(headers_en, 1):
        cell = ws1.cell(row=1, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    headers_cn = [
        "文件名",
        "材料品牌",
        "材料",
        "颜色",
        "数量",
        "打印机",
        "喷嘴直径",
        "层高(mm)",
        "墙层数",
        "填充密度(%)",
    ]
    for col, val in enumerate(headers_cn, 1):
        cell = ws1.cell(row=2, column=col, value=val)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = center_align
        cell.border = thin_border

    brand_examples = user_brands[:3] if len(user_brands) >= 3 else user_brands
    while len(brand_examples) < 3:
        brand_examples.append("Generic")

    examples = [
        ["model1.stl", brand_examples[0], "PLA", "白色", 1, "Bambu Lab A1", 0.4, 0.2, 3, 20],
        ["model2.stl", "", "", "", "", "", "", 0.16, 4, 15],
        [
            "model3.stl",
            brand_examples[2] if len(brand_examples) > 2 else "Generic",
            "PETG",
            "黑色",
            2,
            "Creality K1 Max",
            0.6,
            0.28,
            2,
            10,
        ],
    ]
    for r, row_data in enumerate(examples, 3):
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=col, value=val if val != "" else None)
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = thin_border

    note_row = 6
    note_text = "提示：空白单元格 = 使用系统默认值，填写 = 覆盖默认值。第一行（英文列名）必须保留。"
    ws1.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
    note_cell = ws1.cell(row=note_row, column=1, value=note_text)
    note_cell.font = note_font
    note_cell.fill = note_fill
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for c in range(1, 11):
        ws1.cell(row=note_row, column=c).border = thin_border

    col_widths_s1 = [16, 14, 12, 10, 10, 20, 14, 16, 12, 16]
    for i, w in enumerate(col_widths_s1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("参数说明")

    param_headers = ["参数", "英文列名", "说明", "默认值", "可选值"]
    for col, val in enumerate(param_headers, 1):
        cell = ws2.cell(row=1, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    brands_display = ", ".join(user_brands)
    params = [
        ["文件名", "filename", "必填，与压缩包内模型文件名（不含扩展名）匹配", "—", "—"],
        ["材料品牌", "material_brand", "可选，耗材品牌名称", "Generic", brands_display],
        ["材料", "material_type", "可选，材料类型（如 PETG, PLA, ABS）", "—", "PETG, PLA, PLA+, ABS, ASA, TPU, PA, PC"],
        ["颜色", "color", "可选，模型颜色", "使用表单默认", "白色, 黑色, 红色, 蓝色, 绿色 等"],
        ["数量", "quantity", "可选，正整数", "使用表单默认", "1, 2, 3, ..."],
        ["打印机", "printer", "可选，打印机型号", "使用系统默认", "Bambu Lab A1, Creality K1, Prusa MK4 等"],
        ["喷嘴直径", "nozzle", "可选，单位 mm", "0.4", "0.2, 0.4, 0.6, 0.8"],
        ["层高", "layer_height", "可选，单位 mm", "0.2", "0.08, 0.10, 0.12, 0.16, 0.20, 0.28, 0.32"],
        ["墙层数", "wall_count", "可选，外壁层数", "3", "2, 3, 4, 5, 6, 8"],
        ["填充密度", "infill", "可选，百分比", "20", "5, 10, 15, 20, 25, 30, 40, 50, 60, 80, 100"],
    ]
    for r, row_data in enumerate(params, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=col, value=val)
            cell.font = normal_font
            cell.alignment = left_align if col >= 3 else center_align
            cell.border = thin_border

    col_widths_s2 = [12, 16, 40, 16, 44]
    for i, w in enumerate(col_widths_s2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=zip_import_template.xlsx"},
    )
