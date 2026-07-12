"""Quote history export services (CSV / XLSX / PDF)."""

import csv
import io
import logging
from datetime import datetime
from typing import Optional

from fastapi import Depends, HTTPException, Request
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel

from app.db import get_db_session
from app.deps import get_current_user, get_membership_effective, is_member_user
from app.models_orm import QuoteHistory, User

logger = logging.getLogger(__name__)


_EXPORT_COLUMNS = [
    ("filename", "文件名"),
    ("material", "材料"),
    ("color", "颜色"),
    ("quantity", "数量"),
    ("volume_cm3", "体积(cm³)"),
    ("weight_g", "重量(g)"),
    ("estimated_time_h", "打印时间(h)"),
    ("cost_cny", "成本(元)"),
    ("created_at", "创建时间"),
    ("printer_model", "打印机型号"),
    ("slicer_preset_id", "切片预设ID"),
    ("nozzle_diameter", "喷嘴直径(mm)"),
    ("layer_height", "层高(mm)"),
    ("wall_count", "墙层数"),
    ("infill", "填充(%)"),
    ("brand", "品牌"),
]


def _build_export_query(db, user_id: int, material: Optional[str], date_from: Optional[str], date_to: Optional[str]):
    """Build filtered query for quote history export."""
    query = db.query(QuoteHistory).filter(QuoteHistory.user_id == user_id)
    if material:
        query = query.filter(QuoteHistory.material == material)
    if date_from:
        query = query.filter(QuoteHistory.created_at >= date_from)
    if date_to:
        date_to_end = date_to if "T" in date_to else f"{date_to}T23:59:59"
        query = query.filter(QuoteHistory.created_at <= date_to_end)
    return query.order_by(QuoteHistory.id.desc())


def _row_to_export_list(row) -> list:
    """Convert a QuoteHistory ORM row to an export value list."""
    return [
        row.filename or "",
        row.material or "",
        row.color or "",
        int(row.quantity or 0),
        round(float(row.volume_cm3 or 0), 2),
        round(float(row.weight_g or 0), 2),
        round(float(row.estimated_time_h or 0), 2),
        round(float(row.cost_cny or 0), 2),
        str(row.created_at) if row.created_at else "",
        row.printer_model or "",
        int(row.slicer_preset_id) if row.slicer_preset_id is not None else "",
        round(float(row.nozzle_diameter), 2) if row.nozzle_diameter is not None else "",
        round(float(row.layer_height), 2) if row.layer_height is not None else "",
        int(row.wall_count) if row.wall_count is not None else "",
        int(row.infill) if row.infill is not None else "",
        row.brand or "",
    ]


async def export_quote_history(
    request: Request,
    format: str,
    material: Optional[str],
    date_from: Optional[str],
    date_to: Optional[str],
    current_user: dict = Depends(get_current_user),
):
    """Export quote history as CSV or XLSX file download."""
    if not is_member_user(current_user):
        raise HTTPException(status_code=403, detail="导出功能仅限会员使用")

    fmt = (format or "csv").strip().lower()
    if fmt not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="format 参数必须为 csv 或 xlsx")

    uid = int(current_user["id"])
    logger.info(f"用户 {uid} 请求导出报价历史 format={fmt} material={material} date_from={date_from} date_to={date_to}")

    try:
        with get_db_session() as db:
            query = _build_export_query(db, uid, material, date_from, date_to)
            rows = query.all()
            export_rows = [_row_to_export_list(r) for r in rows]

        if not rows:
            raise HTTPException(status_code=404, detail="没有符合条件的报价记录可导出")

        headers = [col[1] for col in _EXPORT_COLUMNS]
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

        if fmt == "csv":
            buf = io.StringIO()
            buf.write("\ufeff")
            writer = csv.writer(buf)
            writer.writerow(headers)
            for row_values in export_rows:
                writer.writerow(row_values)
            buf.seek(0)
            filename = f"quote_history_{timestamp}.csv"
            return StreamingResponse(
                buf,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        else:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "报价历史"

            Font(bold=True, size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(bold=True, size=11, color="FFFFFF")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col_idx, (_, title) in enumerate(_EXPORT_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=title)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row_idx, row_values in enumerate(export_rows, 2):
                for col_idx, val in enumerate(row_values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border

            for col_idx, (_, title) in enumerate(_EXPORT_COLUMNS, 1):
                max_len = len(title) * 2
                for row in ws.iter_rows(
                    min_row=2, max_row=min(len(export_rows) + 1, 102), min_col=col_idx, max_col=col_idx
                ):
                    for cell in row:
                        if cell.value is not None:
                            val_len = len(str(cell.value))
                            max_len = max(max_len, val_len)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            filename = f"quote_history_{timestamp}.xlsx"
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出报价历史失败: user_id={uid} error={str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


async def export_quote_pdf(
    request: Request,
    ids: str,
    current_user: dict = Depends(get_current_user),
):
    """Export selected quote history items as a branded PDF quote."""
    from app.config import MEMBER_DISCOUNT_PERCENT
    from app.services.pdf import generate_pdf_quote

    if not is_member_user(current_user):
        raise HTTPException(status_code=403, detail="PDF?????????????")

    uid = int(current_user["id"])

    if not ids or not ids.strip():
        raise HTTPException(status_code=400, detail="???????ID (ids??)")

    try:
        id_list = [int(x.strip()) for x in ids.split(",") if x.strip()]
    except ValueError:
        raise HTTPException(status_code=400, detail="ids?????????????????")

    if not id_list:
        raise HTTPException(status_code=400, detail="???????????ID")
    if len(id_list) > 100:
        raise HTTPException(status_code=400, detail="??????100???")

    try:
        with get_db_session() as db:
            user_row = db.query(User).filter(User.id == uid).first()
            brand_name = user_row.brand_name if user_row else ""
            brand_logo_url = user_row.brand_logo_url if user_row else ""
            brand_phone = user_row.brand_phone if user_row else ""
            brand_contact_email = user_row.brand_contact_email if user_row else ""
            brand_address = user_row.brand_address if user_row else ""
            brand_note = user_row.brand_note if user_row else ""

            rows = (
                db.query(QuoteHistory)
                .filter(QuoteHistory.id.in_(id_list), QuoteHistory.user_id == uid)
                .order_by(QuoteHistory.id.asc())
                .all()
            )
            items = []
            for r in rows:
                items.append(
                    {
                        "filename": r.filename or "",
                        "material": r.material or "",
                        "color": r.color or "",
                        "quantity": int(r.quantity or 1),
                        "volume_cm3": round(float(r.volume_cm3 or 0), 2),
                        "weight_g": round(float(r.weight_g or 0), 2),
                        "estimated_time_h": round(float(r.estimated_time_h or 0), 2),
                        "cost_cny": round(float(r.cost_cny or 0), 2),
                        "printer_model": r.printer_model or "",
                        "nozzle_diameter": round(float(r.nozzle_diameter), 2) if r.nozzle_diameter is not None else "",
                        "layer_height": round(float(r.layer_height), 2) if r.layer_height is not None else 0,
                        "wall_count": int(r.wall_count) if r.wall_count is not None else 0,
                        "infill_percent": int(r.infill) if r.infill is not None else 0,
                        "brand": r.brand or "",
                    }
                )

        if not items:
            raise HTTPException(status_code=404, detail="??????????")

        membership_level, _ = get_membership_effective(current_user)
        discount_pct = 0.0
        if membership_level == "member":
            discount_pct = float(MEMBER_DISCOUNT_PERCENT or 0.0)
            discount_pct = max(0.0, min(90.0, discount_pct))

        pdf_bytes = generate_pdf_quote(
            items=items,
            brand_name=brand_name or "",
            brand_logo_url=brand_logo_url or "",
            brand_phone=brand_phone or "",
            brand_contact_email=brand_contact_email or "",
            brand_address=brand_address or "",
            brand_note=brand_note or "",
            member_discount_percent=discount_pct,
        )

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"quote_{timestamp}.pdf"

        logger.info(f"?? {uid} ??PDF?????? {len(items)} ???")
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(pdf_bytes)),
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"??PDF?????: user_id={uid} error={str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"PDF????: {str(e)}")


class PdfInlineItem(BaseModel):
    filename: str = ""
    material: str = ""
    color: str = ""
    quantity: int = 1
    volume_cm3: float = 0
    weight_g: float = 0
    estimated_time_h: float = 0
    cost_cny: float = 0
    created_at: str = ""
    status: str = "success"
    printer_model: str = ""
    nozzle_diameter: str = ""
    layer_height: float = 0
    wall_count: int = 0
    infill_percent: int = 0
    brand: str = ""
    thumbnail_b64: str = ""


class PdfInlineRequest(BaseModel):
    items: list


async def export_pdf_inline(
    payload: PdfInlineRequest,
    request: Request,
    current_user: dict = Depends(get_current_user),
):
    """POST /api/quote/export-pdf-inline — 从当前页面结果直接生成PDF（不查DB）"""
    from app.config import MEMBER_DISCOUNT_PERCENT
    from app.services.pdf import generate_pdf_quote

    if not is_member_user(current_user):
        raise HTTPException(status_code=403, detail="会员专属功能")

    if not payload.items:
        raise HTTPException(status_code=400, detail="没有可导出的项目")

    uid = int(current_user["id"])
    brand_name = brand_logo_url = brand_phone = brand_contact_email = brand_address = brand_note = ""
    with get_db_session() as db:
        user = db.query(User).filter(User.id == uid).first()
        if user:
            brand_name = user.brand_name or ""
            brand_logo_url = user.brand_logo_url or ""
            brand_phone = user.brand_phone or ""
            brand_contact_email = user.brand_contact_email or ""
            brand_address = user.brand_address or ""
            brand_note = user.brand_note or ""

    level, _ = get_membership_effective(current_user)
    discount_pct = float(MEMBER_DISCOUNT_PERCENT or 0.0) if level == "member" else 0.0
    discount_pct = max(0.0, min(90.0, discount_pct))

    items = [item.model_dump() for item in payload.items]
    pdf_bytes = generate_pdf_quote(
        items=items,
        brand_name=brand_name,
        brand_logo_url=brand_logo_url,
        brand_phone=brand_phone,
        brand_contact_email=brand_contact_email,
        brand_address=brand_address,
        brand_note=brand_note,
        member_discount_percent=discount_pct,
    )

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"quote_{timestamp}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
        },
    )
