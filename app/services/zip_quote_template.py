"""Excel template generation for ZIP checklist imports."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TEMPLATE_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
TEMPLATE_FILENAME = "zip_import_template.xlsx"


def build_zip_template_bytes(user_brands: list[str]) -> bytes:
    """Build the import workbook while keeping brand data caller-controlled."""
    brands = list(user_brands or [])
    if not brands:
        brands = ["eSUN", "Generic", "Hatchbox", "Polymaker", "Sunlu"]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    sub_header_font = Font(name="Microsoft YaHei", bold=True, size=11)
    sub_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    normal_font = Font(name="Microsoft YaHei", size=11)
    note_font = Font(name="Microsoft YaHei", size=10, color="666666")
    note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "导入模板"

    headers_en = [
        "filename",
        "material_brand",
        "material_type",
        "color",
        "quantity",
        "printer",
        "nozzle",
        "layer_height",
        "wall_count",
        "infill",
    ]
    for column, value in enumerate(headers_en, 1):
        cell = worksheet.cell(row=1, column=column, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    headers_cn = [
        "文件名",
        "材料品牌",
        "材料",
        "颜色",
        "数量",
        "打印机",
        "喷嘴直径",
        "层高(mm)",
        "墙层数",
        "填充密度(%)",
    ]
    for column, value in enumerate(headers_cn, 1):
        cell = worksheet.cell(row=2, column=column, value=value)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = center_align
        cell.border = thin_border

    brand_examples = brands[:3]
    while len(brand_examples) < 3:
        brand_examples.append("Generic")
    examples = [
        ["model1.stl", brand_examples[0], "PLA", "白色", 1, "Bambu Lab A1", 0.4, 0.2, 3, 20],
        ["model2.stl", "", "", "", "", "", "", 0.16, 4, 15],
        [
            "model3.stl",
            brand_examples[2],
            "PETG",
            "黑色",
            2,
            "Creality K1 Max",
            0.6,
            0.28,
            2,
            10,
        ],
    ]
    for row, row_data in enumerate(examples, 3):
        for column, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row, column=column, value=value if value != "" else None)
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = thin_border

    note_row = 6
    worksheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
    note_cell = worksheet.cell(
        row=note_row,
        column=1,
        value="提示：空白单元格 = 使用系统默认值，填写 = 覆盖默认值。第一行（英文列名）必须保留。",
    )
    note_cell.font = note_font
    note_cell.fill = note_fill
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for column in range(1, 11):
        worksheet.cell(row=note_row, column=column).border = thin_border

    for column, width in enumerate([16, 14, 12, 10, 10, 20, 14, 16, 12, 16], 1):
        worksheet.column_dimensions[get_column_letter(column)].width = width

    parameters = workbook.create_sheet("参数说明")
    parameter_headers = ["参数", "英文列名", "说明", "默认值", "可选值"]
    for column, value in enumerate(parameter_headers, 1):
        cell = parameters.cell(row=1, column=column, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    params = [
        ["文件名", "filename", "必填，与压缩包内模型文件名（不含扩展名）匹配", "—", "—"],
        ["材料品牌", "material_brand", "可选，耗材品牌名称", "Generic", ", ".join(brands)],
        ["材料", "material_type", "可选，材料类型（如 PETG, PLA, ABS）", "—", "PETG, PLA, PLA+, ABS, ASA, TPU, PA, PC"],
        ["颜色", "color", "可选，模型颜色", "使用表单默认", "白色, 黑色, 红色, 蓝色, 绿色 等"],
        ["数量", "quantity", "可选，正整数", "使用表单默认", "1, 2, 3, ..."],
        ["打印机", "printer", "可选，打印机型号", "使用系统默认", "Bambu Lab A1, Creality K1, Prusa MK4 等"],
        ["喷嘴直径", "nozzle", "可选，单位 mm", "0.4", "0.2, 0.4, 0.6, 0.8"],
        ["层高", "layer_height", "可选，单位 mm", "0.2", "0.08, 0.10, 0.12, 0.16, 0.20, 0.28, 0.32"],
        ["墙层数", "wall_count", "可选，外壁层数", "3", "2, 3, 4, 5, 6, 8"],
        ["填充密度", "infill", "可选，百分比", "20", "5, 10, 15, 20, 25, 30, 40, 50, 60, 80, 100"],
    ]
    for row, row_data in enumerate(params, 2):
        for column, value in enumerate(row_data, 1):
            cell = parameters.cell(row=row, column=column, value=value)
            cell.font = normal_font
            cell.alignment = left_align if column >= 3 else center_align
            cell.border = thin_border

    for column, width in enumerate([12, 16, 40, 16, 44], 1):
        parameters.column_dimensions[get_column_letter(column)].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
