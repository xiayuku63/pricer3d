"""
ZIP upload Excel checklist parsing and model matching logic.

Extracted from routes_zip_quote.py to keep route file focused on HTTP concerns.
"""

import io
import os
import logging
from typing import List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Fuzzy header patterns: canonical_key → list of matching patterns
HEADER_PATTERNS = {
    'filename': ['filename', '文件名', '文件', '名称', '模型名', 'model', 'name', 'stl'],
    'printer_model': ['printer', '打印机', '机型', '打印机型号', 'printer_model'],
    'nozzle': ['nozzle', '喷嘴', '喷嘴直径', '口径'],
    'layer_height': ['layer_height', '层高', 'layer height', '层厚', '层高mm'],
    'wall_count': ['wall_count', 'wall', '墙层数', '壁数', '外墙', 'perimeters', 'perimeter', '墙数'],
    'infill': ['infill', 'infill_density', '填充', '填充率', '填充密度', 'density', 'infill%'],
    'material_brand': ['material_brand', '材料品牌', '品牌', 'brand', '材料牌号'],
    'material_type': ['material', '材料', 'material_type', '材料类型', '材质'],
    'color': ['color', '颜色', 'colour', '色彩'],
    'quantity': ['quantity', 'qty', '数量', '数目', '件数', 'count', '个数'],
}


def _match_headers(headers: list) -> dict:
    """Fuzzy-match header strings to canonical keys.

    Returns {canonical_key: col_index} for matched columns.
    Matching priority: exact > contains-pattern > pattern-contains-header.
    """
    result = {}
    used_keys = set()

    for idx, raw in enumerate(headers):
        h = str(raw).strip().lower()
        if not h:
            continue

        # Pass 1: exact match
        matched_key = None
        for key, patterns in HEADER_PATTERNS.items():
            if key in used_keys:
                continue
            if h in patterns:
                matched_key = key
                break

        # Pass 2: pattern contained in header text
        if not matched_key:
            for key, patterns in HEADER_PATTERNS.items():
                if key in used_keys:
                    continue
                if any(p in h for p in patterns):
                    matched_key = key
                    break

        # Pass 3: header text contained in any pattern
        if not matched_key:
            for key, patterns in HEADER_PATTERNS.items():
                if key in used_keys:
                    continue
                if any(h in p for p in patterns):
                    matched_key = key
                    break

        if matched_key:
            result[matched_key] = idx
            used_keys.add(matched_key)

    return result


# Valid ranges for checklist parameters
VALID_RANGES = {
    "nozzle": ["0.2", "0.4", "0.6", "0.8"],
    "wall_count": [2, 3, 4, 5, 6, 8],
    "infill": [5, 10, 15, 20, 25, 30, 40, 50, 60, 80, 100],
}

# Default values for invalid parameters
_DEFAULT_VALUES = {
    "nozzle": "0.4",
    "layer_height": 0.2,
    "wall_count": 3,
    "infill": 20,
}

# Layer height valid values per nozzle diameter
LAYER_HEIGHT_BY_NOZZLE = {
    '0.2': {'min': 0.06, 'max': 0.14, 'valid': [0.06, 0.08, 0.10, 0.12, 0.14], 'default': 0.10},
    '0.4': {'min': 0.08, 'max': 0.28, 'valid': [0.08, 0.12, 0.16, 0.20, 0.24, 0.28], 'default': 0.20},
    '0.6': {'min': 0.18, 'max': 0.42, 'valid': [0.18, 0.24, 0.30, 0.36, 0.42], 'default': 0.30},
    '0.8': {'min': 0.24, 'max': 0.56, 'valid': [0.24, 0.32, 0.40, 0.48, 0.56], 'default': 0.40},
}


def _validate_checklist_item(item: dict, filename_stem: str) -> List[dict]:
    """Validate a single checklist item against known valid ranges.
    Returns a list of warning dicts: {filename, param, value, default_used}."""
    warnings = []

    # Resolve effective nozzle for this item (checklist value or default)
    nozzle_raw = str(item.get("nozzle", "")).strip()
    nozzle_key = nozzle_raw if nozzle_raw in LAYER_HEIGHT_BY_NOZZLE else _DEFAULT_VALUES["nozzle"]
    nozzle_settings = LAYER_HEIGHT_BY_NOZZLE.get(nozzle_key, LAYER_HEIGHT_BY_NOZZLE[_DEFAULT_VALUES["nozzle"]])

    # Validate layer_height against per-nozzle valid values
    lh_parsed_key = "layer_height_parsed"
    lh_raw = item.get("layer_height", "")
    lh_val = item.get(lh_parsed_key, None)
    if lh_val is None and lh_raw:
        try:
            lh_val = float(lh_raw)
        except (ValueError, TypeError):
            lh_val = None
    if lh_raw and lh_val is not None:
        valid_heights = nozzle_settings["valid"]
        if lh_val not in valid_heights:
            # Find closest valid value
            closest = min(valid_heights, key=lambda v: abs(v - lh_val))
            warnings.append({
                "filename": filename_stem,
                "param": "layer_height",
                "value": str(lh_val),
                "default_used": str(closest),
                "reason": f"层高{lh_val}mm不是喷嘴{nozzle_key}mm的有效值，已调整为最近的{closest}mm",
            })
            item[lh_parsed_key] = closest

    # Validate numeric/string params in VALID_RANGES
    for param, valid_values in VALID_RANGES.items():
        # Determine parsed value
        parsed_key = param + "_parsed"
        raw = item.get(param, "")

        if param in ("wall_count", "infill"):
            val = item.get(parsed_key, None)
            if val is None and raw:
                try:
                    val = int(float(raw))
                except (ValueError, TypeError):
                    val = None
            if raw and val is not None and val not in valid_values:
                default = _DEFAULT_VALUES[param]
                warnings.append({
                    "filename": filename_stem,
                    "param": param,
                    "value": str(val),
                    "default_used": str(default),
                })
                item[parsed_key] = default
        elif param == "nozzle":
            if raw and str(raw).strip() not in valid_values:
                default = _DEFAULT_VALUES[param]
                warnings.append({
                    "filename": filename_stem,
                    "param": param,
                    "value": str(raw),
                    "default_used": str(default),
                })
                item[param] = default

    # Validate printer_model
    printer_name = str(item.get("printer_model", "")).strip()
    if printer_name:
        from app.printers import PRINTER_MODELS
        name_lower = printer_name.lower()
        found = any(pm.get("name", "").lower() == name_lower for pm in PRINTER_MODELS)
        if not found:
            warnings.append({
                "filename": filename_stem,
                "param": "printer_model",
                "value": printer_name,
                "default_used": "系统默认",
            })

    # Validate material_brand (warn if provided but not recognized)
    brand = str(item.get("material_brand", "")).strip()
    if brand:
        from .config import DEFAULT_MATERIALS
        known_brands = {m.upper() for m in DEFAULT_MATERIALS if isinstance(m, str)}
        # Also accept common brand names
        known_brand_names = {
            "GENERIC", "ESUN", "E-SUN", "HATCHBOX", "POLYMAKER", "PRUSAMENT",
            "SUNLU", "ERYONE", "OVERTURE", "INLAND", "MAKERBOT", "FILATECH",
            "ELEGOO", "ANYCUBIC", "FLASHFORGE", "BAMBU LAB", "QIDI", "CREALITY",
        }
        if brand.upper() not in known_brands and brand.upper() not in known_brand_names:
            warnings.append({
                "filename": filename_stem,
                "param": "material_brand",
                "value": brand,
                "default_used": "Generic",
            })

    return warnings


def _collect_all_warnings(checklist: Optional[List[dict]]) -> List[dict]:
    """Collect all _warnings from checklist items."""
    if not checklist:
        return []
    warnings = []
    for item in checklist:
        w = item.get("_warnings")
        if w:
            warnings.extend(w)
    return warnings


def _parse_excel_checklist(file_bytes: bytes) -> Optional[List[dict]]:
    """Parse Excel checklist. Returns list of {filename, material_brand, material_type, printer_model, nozzle, layer_height, wall_count, infill}
    or None if no checklist found / parse error."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"Failed to open Excel workbook: {e}")
        return None

    ws = wb.active
    if not ws:
        return None

    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1000, 1000), values_only=True))
    if len(rows) < 2:
        return None

    # Find the header row — scan rows until we find one with a 'filename'-mapped column
    header_row_idx = None
    col_map = {}
    for row_idx, row in enumerate(rows):
        header = [str(c).strip().lower() if c else "" for c in row]
        test_map = _match_headers(header)
        if "filename" in test_map:
            header_row_idx = row_idx
            col_map = test_map
            break

    if header_row_idx is None:
        logger.warning("Excel has no 'filename' column — not a valid checklist")
        return None

    items = []
    all_warnings = []
    for row in rows[header_row_idx + 1:]:
        item = {}
        for key, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                item[key] = str(val).strip()
            else:
                item[key] = ""

        # Skip rows with no filename
        if not item.get("filename"):
            continue

        # Normalize filename — strip extension and any subfolder path
        fn = item["filename"]
        # Strip subfolder path (e.g. "DB4打印件/model.stl" → "model.stl")
        fn = os.path.basename(fn)
        dot_pos = fn.rfind(".")
        if dot_pos > 0:
            fn = fn[:dot_pos]
        item["filename_stem"] = fn.lower()

        # Parse numeric fields with validation
        for num_key in ("layer_height", "wall_count", "infill"):
            val = item.get(num_key, "")
            if val == "":
                continue
            try:
                if num_key == "layer_height":
                    item[num_key + "_parsed"] = float(val)
                elif num_key == "wall_count":
                    item[num_key + "_parsed"] = int(float(val))
                else:
                    item[num_key + "_parsed"] = int(float(val.replace("%", "")))
            except (ValueError, TypeError):
                pass

        # Parse quantity (positive integer)
        qty_raw = item.get("quantity", "")
        if qty_raw:
            try:
                qty_val = int(float(qty_raw))
                if qty_val > 0:
                    item["quantity_parsed"] = qty_val
            except (ValueError, TypeError):
                pass

        # Validate parameters against known ranges
        item_warnings = _validate_checklist_item(item, item.get("filename_stem", ""))
        if item_warnings:
            item["_warnings"] = item_warnings
            all_warnings.extend(item_warnings)

        items.append(item)

    wb.close()
    return items if items else None


def _match_checklist_to_models(
    checklist: List[dict],
    stl_files: List[dict],  # [{filename, name_stem, file_bytes}]
) -> dict:
    """
    Match checklist items to STL files by filename stem.
    Returns:
        {
            matched: [{checklist_item, stl_info}],      # full match
            checklist_only: [{...}],                    # in checklist but no STL
            stl_only: [{...}],                          # STL files with no checklist entry
            match_mode: "all" | "partial" | "none",
        }
    """
    # Build STL stem → info map
    stl_by_stem = {}
    for f in stl_files:
        stl_by_stem[f["name_stem"]] = f

    matched_stems = set()
    matched = []
    checklist_only = []

    for item in checklist:
        stem = item.get("filename_stem", "")
        if stem and stem in stl_by_stem:
            matched.append({
                "checklist": item,
                "stl": stl_by_stem[stem],
            })
            matched_stems.add(stem)
        else:
            checklist_only.append(item)

    # STL files not in checklist
    stl_only = [f for f in stl_files if f["name_stem"] not in matched_stems]

    # Determine match mode
    if len(matched) == len(stl_files) and len(checklist_only) == 0 and len(stl_only) == 0:
        match_mode = "all"
    elif len(matched) == 0:
        match_mode = "none"
    else:
        match_mode = "partial"

    return {
        "matched": matched,
        "checklist_only": checklist_only,
        "stl_only": stl_only,
        "match_mode": match_mode,
    }
