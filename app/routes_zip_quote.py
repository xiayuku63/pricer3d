"""
ZIP upload quote route — parse Excel checklist + STL models from .zip archive.

Excel checklist columns (header row, fuzzy-matched, case-insensitive):
    filename | material_brand | material_type | color | quantity | printer | nozzle | layer_height | wall_count | infill
    文件名   | 材料品牌       | 材料          | 颜色  | 数量     | 打印机  | 喷嘴直径 | 层高         | 墙层数     | 填充密度

Matching: compare checklist filename stem with STL filename stem (case-insensitive,
ignoring extension). Reports full / partial / none match status.
"""

import io
import json
import logging
import os
import uuid
import zipfile
from typing import List, Optional

from fastapi import Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import (
    DEFAULT_MATERIALS,
    DEFAULT_PRICING_CONFIG,
    MAX_FILES_PER_REQUEST,
    MAX_FILE_SIZE_BYTES,
    SUPPORTED_EXTENSIONS,
    QUOTE_CONCURRENCY,
)
from .deps import get_current_user, get_membership_effective
from .audit import write_audit_event
from .zip_parser import (
    HEADER_PATTERNS,
    VALID_RANGES,
    _DEFAULT_VALUES,
    LAYER_HEIGHT_BY_NOZZLE,
    _match_headers,
    _validate_checklist_item,
    _collect_all_warnings,
    _parse_excel_checklist,
    _match_checklist_to_models,
)

logger = logging.getLogger(__name__)

# Max ZIP file size (1GB)
MAX_ZIP_SIZE_BYTES = 1024 * 1024 * 1024


async def zip_quote(
    request: Request,
    file: UploadFile = File(...),
    material: str = Form("PLA"),
    color: str = Form("White"),
    quantity: int = Form(1),
    printer_model: Optional[str] = Form(default=None),
    slicer_preset_id: Optional[int] = Form(default=None),
    current_user=Depends(get_current_user),
):
    """
    Upload a .zip file containing an Excel checklist (.xlsx/.xls) + STL model files.
    Returns match analysis + triggers slicing for matched models.

    Request: multipart/form-data with:
        file: the .zip file (required)
        material: default material (optional, default PLA)
        color: default color (optional, default White)
        quantity: default quantity (optional, default 1)
    """
    import asyncio

    try:
        # Validate file extension
        fname = (file.filename or "").lower()
        if not fname.endswith(".zip"):
            raise HTTPException(status_code=400, detail="请上传 .zip 压缩文件")

        # Read file content
        content = await file.read()
        if len(content) >= MAX_ZIP_SIZE_BYTES:
            raise HTTPException(status_code=400, detail=f"ZIP 文件必须小于 {MAX_ZIP_SIZE_BYTES // (1024*1024)}MB")

        # Parse ZIP
        try:
            zf = zipfile.ZipFile(io.BytesIO(content))
        except zipfile.BadZipFile:
            raise HTTPException(status_code=400, detail="无效的 ZIP 文件，请检查文件完整性")

        # Separate Excel and STL files from ZIP
        excel_bytes = None
        stl_files = []  # [{filename, name_stem, file_bytes}]

        for entry in zf.infolist():
            # Skip directories and __MACOSX junk
            if entry.is_dir():
                continue
            bn = os.path.basename(entry.filename)
            if bn.startswith(".") or bn.startswith("__MACOSX") or bn.startswith("~$"):
                continue

            entry_bytes = zf.read(entry)

            lower = bn.lower()
            if lower.endswith((".xlsx", ".xls")):
                if excel_bytes is None:
                    excel_bytes = entry_bytes
                    logger.info(f"ZIP: found Excel checklist: {bn}")
            else:
                ext = os.path.splitext(lower)[1]
                if ext in SUPPORTED_EXTENSIONS:
                    stem = os.path.splitext(bn)[0].lower()
                    stl_files.append({
                        "filename": bn,
                        "name_stem": stem,
                        "file_bytes": entry_bytes,
                        "ext": ext,
                    })
                    logger.info(f"ZIP: found model file: {bn} (stem={stem})")

        zf.close()

        if not stl_files:
            raise HTTPException(status_code=400, detail="ZIP 中未找到支持的模型文件（.stl/.stp/.step/.obj/.3mf）")

        if len(stl_files) > MAX_FILES_PER_REQUEST:
            raise HTTPException(status_code=400, detail=f"ZIP 中模型文件数量不能超过 {MAX_FILES_PER_REQUEST} 个")

        # 免费用户累计模型总数限制
        from .deps import is_member_user
        from .db import get_db_session
        from sqlalchemy import func as sqlfunc
        from .models_orm import QuoteHistory
        from .config import FREE_TOTAL_MODEL_LIMIT
        if not is_member_user(current_user):
            with get_db_session() as db:
                existing_count = db.query(sqlfunc.count(QuoteHistory.id)).filter(
                    QuoteHistory.user_id == current_user["id"],
                    QuoteHistory.status == "success",
                ).scalar() or 0
            if existing_count >= FREE_TOTAL_MODEL_LIMIT:
                raise HTTPException(status_code=400, detail=f"免费用户最多累计 {FREE_TOTAL_MODEL_LIMIT} 个模型，升级会员无限制")

        # Parse Excel checklist
        checklist = _parse_excel_checklist(excel_bytes) if excel_bytes else None

        # Pre-save all model files to disk so thumbnails work even if slicing fails
        from app.utils import _user_base_dir
        _user_folder = f"user_{current_user['id']}_{current_user['username']}"
        _zip_job_id = uuid.uuid4().hex[:8]
        _zip_uploads_dir = os.path.join(_user_base_dir(), _user_folder, "uploads", _zip_job_id)
        os.makedirs(_zip_uploads_dir, exist_ok=True)
        for _sf in stl_files:
            _saved = os.path.join(_zip_uploads_dir, _sf["filename"])
            with open(_saved, "wb") as _f:
                _f.write(_sf["file_bytes"])
            _sf["_pre_saved_path"] = _saved

        # Match checklist to models
        if checklist:
            match_result = _match_checklist_to_models(checklist, stl_files)
        else:
            match_result = {
                "matched": [],
                "checklist_only": [],
                "stl_only": stl_files,
                "match_mode": "none",
            }

        # Get user materials + pricing + defaults
        from .db import get_db_session
        from .models_orm import User as UserORM
        with get_db_session() as db:
            u = db.query(UserORM).filter(UserORM.id == current_user["id"]).first()
            user_materials = json.loads(u.materials) if u and u.materials else DEFAULT_MATERIALS
            pricing_config = json.loads(u.pricing_config) if u and u.pricing_config else DEFAULT_PRICING_CONFIG
            default_printer_id = u.default_printer_id if u else None
            default_nozzle = u.default_nozzle if u else None
            default_slicer_preset_id = u.default_slicer_preset_id if u else None

        # Build match status message
        total_stl = len(stl_files)
        if match_result["match_mode"] == "all":
            match_msg = f"✅ 全部模型预设生效（{len(match_result['matched'])}/{total_stl} 个文件匹配）"
        elif match_result["match_mode"] == "partial":
            match_msg = f"⚠️ 部分模型预设生效，请检查清单（{len(match_result['matched'])} 匹配 / {len(match_result['checklist_only'])} 清单多余 / {len(match_result['stl_only'])} 无预设）"
        else:
            if checklist:
                match_msg = f"❌ 全部模型预设未生效，请检查清单（{len(stl_files)} 个模型均未匹配）"
            else:
                match_msg = "ℹ️ 未包含 Excel 清单，使用默认参数"

        # Now process matched + stl_only files via SSE streaming
        from calculator.cost import process_single_file
        from .slicer_presets import get_slicer_preset_by_id

        # Resolve printer name + nozzle from checklist -> compound_id
        from app.printers import PRINTER_MODELS, resolve_printer
        def _lookup_printer(printer_name, nozzle_str):
            """Return compound_id (e.g. bambu_a1_04) or None if not found."""
            if not printer_name or not str(printer_name).strip():
                return None
            name_lower = str(printer_name).strip().lower()
            for pm in PRINTER_MODELS:
                if pm.get("name", "").lower() == name_lower:
                    pid = pm["id"]
                    nz = None
                    if nozzle_str and str(nozzle_str).strip():
                        try:
                            nz = float(str(nozzle_str).strip())
                        except (ValueError, TypeError):
                            pass
                    resolved = resolve_printer(pid, nz)
                    if resolved:
                        return resolved.get("_compound_id") or pid
                    return pid
            return None

        # Resolve printer for stl_only files: request param > DB default
        _default_compound_id = None
        _default_preset = None
        if printer_model:
            from app.printers import resolve_printer as _rp2
            _resolved = _rp2(printer_model)
            if _resolved:
                _default_compound_id = _resolved.get("_compound_id") or printer_model
        elif default_printer_id:
            nz = float(default_nozzle) if default_nozzle else None
            resolved = resolve_printer(default_printer_id, nz)
            if resolved:
                _default_compound_id = resolved.get("_compound_id") or default_printer_id
        _effective_preset_id = slicer_preset_id if slicer_preset_id is not None else default_slicer_preset_id
        if _effective_preset_id:
            _default_preset = get_slicer_preset_by_id(int(current_user["id"]), _effective_preset_id)

        # Build flat list of files to process
        _files_to_process = []
        for m in match_result["matched"]:
            _files_to_process.append(("matched", m))
        for stl in match_result["stl_only"]:
            _files_to_process.append(("stl_only", stl))
        _total_files = len(_files_to_process)

        async def _generate():
            """Generator that processes each file and yields SSE progress events."""
            results = []

            for _idx, (_file_type, _item) in enumerate(_files_to_process, 1):
                # ── Cancellation detection ──
                if await request.is_disconnected():
                    logger.info("ZIP processing cancelled by client")
                    yield f'data: {json.dumps({"type": "cancelled", "processed": _idx})}\n\n'
                    return
                try:
                    if _file_type == "matched":
                        cl = _item["checklist"]
                        stl = _item["stl"]

                        # Build per-file options from checklist
                        lh = cl.get("layer_height_parsed", 0.2)
                        if isinstance(lh, str):
                            try:
                                lh = float(lh)
                            except (ValueError, TypeError):
                                lh = 0.2
                        wc = cl.get("wall_count_parsed", 3)
                        if isinstance(wc, str):
                            try:
                                wc = int(float(wc))
                            except (ValueError, TypeError):
                                wc = 3
                        inf = cl.get("infill_parsed", 20)
                        if isinstance(inf, str):
                            try:
                                inf = int(float(inf))
                            except (ValueError, TypeError):
                                inf = 20

                        # Use checklist quantity/color/material if present, else form defaults
                        cl_qty = cl.get("quantity_parsed", quantity)
                        cl_color = cl.get("color", "").strip() or color
                        cl_material = cl.get("material_type", "").strip() or material
                        cl_printer = str(cl.get("printer_model", "")).strip()
                        cl_nozzle = str(cl.get("nozzle", "")).strip()
                        compound_id = _lookup_printer(cl_printer, cl_nozzle)

                        file_pricing = dict(pricing_config)
                        if compound_id:
                            file_pricing["printer_model"] = compound_id

                        fake_file = UploadFile(
                            filename=stl["filename"],
                            file=io.BytesIO(stl["file_bytes"]),
                        )

                        result = await process_single_file(
                            fake_file,
                            material=cl_material,
                            layer_height=lh,
                            infill=inf,
                            quantity=cl_qty,
                            color=cl_color,
                            user_materials=user_materials,
                            pricing_config=file_pricing,
                            slicer_preset=None,
                            perimeters=wc,
                            current_user=current_user,
                            auto_orient=False,
                        )
                        result["_checklist_params"] = True
                        result["_checklist_source"] = {
                            "layer_height": lh,
                            "wall_count": wc,
                            "infill": inf,
                            "printer_model": compound_id or cl_printer or "",
                            "nozzle": cl_nozzle or "",
                            "material_type": cl.get("material_type", ""),
                            "material_brand": cl.get("material_brand", ""),
                            "color": cl_color,
                            "quantity": cl_qty,
                        }
                        if result.get("_saved_path"):
                            result["checklist_file_path"] = result["_saved_path"]

                        filename = stl["filename"]
                        pre_saved = stl.get("_pre_saved_path")

                    else:  # stl_only
                        stl = _item
                        fake_file = UploadFile(
                            filename=stl["filename"],
                            file=io.BytesIO(stl["file_bytes"]),
                        )
                        file_pricing = dict(pricing_config)
                        if _default_compound_id:
                            file_pricing["printer_model"] = _default_compound_id

                        result = await process_single_file(
                            fake_file,
                            material=material,
                            layer_height=0.2,
                            infill=20,
                            quantity=quantity,
                            color=color,
                            user_materials=user_materials,
                            pricing_config=file_pricing,
                            slicer_preset=_default_preset,
                            perimeters=3,
                            current_user=current_user,
                            auto_orient=False,
                        )
                        result["_checklist_params"] = False
                        if result.get("_saved_path"):
                            result["checklist_file_path"] = result["_saved_path"]

                        filename = stl["filename"]
                        pre_saved = stl.get("_pre_saved_path")

                    # Always include file path for thumbnail generation
                    if not result.get("checklist_file_path") and pre_saved:
                        result["checklist_file_path"] = pre_saved

                    results.append(result)

                    status = "success" if result.get("status") == "success" else "failed"
                    yield f'data: {json.dumps({"type": "progress", "current": _idx, "total": _total_files, "filename": filename, "status": status})}\n\n'

                except Exception as e:
                    filename = "unknown"
                    pre_saved = None
                    if isinstance(_item, dict):
                        if _file_type == "matched":
                            filename = _item.get("stl", {}).get("filename", "unknown")
                            pre_saved = _item.get("stl", {}).get("_pre_saved_path")
                        else:
                            filename = _item.get("filename", "unknown")
                            pre_saved = _item.get("_pre_saved_path")
                    result = {
                        "filename": filename,
                        "status": "failed",
                        "error": str(e),
                        "cost_cny": 0,
                        "weight_g": 0,
                        "estimated_time_h": 0,
                    }
                    if pre_saved:
                        result["checklist_file_path"] = pre_saved
                    results.append(result)
                    yield f'data: {json.dumps({"type": "progress", "current": _idx, "total": _total_files, "filename": filename, "status": "failed"})}\n\n'

            # Build final payload
            success_items = [r for r in results if r.get("status") == "success"]
            failed_items = [r for r in results if r.get("status") == "failed"]

            payload = {
                "total_files": len(results),
                "success_count": len(success_items),
                "failed_count": len(failed_items),
                "summary_total_cost_cny": round(sum(r.get("cost_cny", 0) for r in success_items), 2),
                "summary_total_weight_g": round(sum(r.get("weight_g", 0) for r in success_items), 2),
                "summary_total_time_h": round(sum(r.get("estimated_time_h", 0) for r in success_items), 2),
                "results": results,
                "match_status": {
                    "mode": match_result["match_mode"],
                    "message": match_msg,
                    "matched_count": len(match_result["matched"]),
                    "checklist_only_count": len(match_result["checklist_only"]),
                    "stl_only_count": len(match_result["stl_only"]),
                    "checklist_only_files": [c.get("filename", c.get("filename_stem", "")) for c in match_result["checklist_only"]],
                    "warnings": _collect_all_warnings(checklist),
                },
            }

            # Save quote history
            from .routes_quote import _save_quote_history
            _save_quote_history(int(current_user["id"]), results)

            write_audit_event(
                action="quote.zip_upload",
                request=request,
                user=current_user,
                detail={
                    "files": len(results),
                    "success": len(success_items),
                    "failed": len(failed_items),
                    "match_mode": match_result["match_mode"],
                    "material": material,
                    "quantity": quantity,
                },
            )

            yield f'data: {json.dumps({"type": "done", **payload})}\n\n'

        return StreamingResponse(_generate(), media_type="text/event-stream")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"ZIP quote failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"INTERNAL_ERROR: ZIP 报价失败 ({str(e)})")


async def download_zip_model(
    file_path: str,
    current_user=Depends(get_current_user),
):
    """
    Download a model file that was saved during ZIP processing.
    Used by frontend to rebuild File objects for preview/re-quote.
    """
    from fastapi.responses import FileResponse
    import os as _os

    # Security: only allow paths under the user's upload directory
    from app.utils import _user_base_dir
    user_folder = f"user_{current_user['id']}_{current_user['username']}"
    allowed_prefix = _os.path.join(_user_base_dir(), user_folder, "uploads")

    abs_path = _os.path.abspath(file_path)
    if not abs_path.startswith(allowed_prefix):
        raise HTTPException(status_code=403, detail="Access denied")

    if not _os.path.isfile(abs_path):
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(abs_path, filename=_os.path.basename(abs_path))


async def download_zip_template(request: Request):
    """Generate and return an xlsx template for ZIP import checklist.

    If user is authenticated, uses their configured material brands in the template.
    Otherwise uses hardcoded defaults.
    """

    # ── Resolve user brands (optional auth) ──
    user_brands = None
    authorization = request.headers.get("authorization")
    if authorization:
        try:
            from jose import jwt as jose_jwt
            from .config import JWT_SECRET_KEY, JWT_ALGORITHM
            from .db import get_db_session
            from .models_orm import User as _UserORM

            token = authorization.replace("Bearer ", "")
            payload = jose_jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
            user_id = int(payload.get("sub", "0"))
            if user_id > 0:
                with get_db_session() as db:
                    u = db.query(_UserORM.materials).filter(_UserORM.id == user_id).first()
                if u and u.materials:
                    materials = json.loads(u.materials)
                    brands = sorted({
                        m.get("brand", "Generic")
                        for m in materials
                        if isinstance(m, dict) and m.get("brand")
                    })
                    if brands:
                        user_brands = brands
        except Exception:
            pass  # fall back to defaults

    # Default brands if not authenticated or no user materials
    if not user_brands:
        user_brands = ["eSUN", "Generic", "Hatchbox", "Polymaker", "Sunlu"]

    # ── Styles ──
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    sub_header_font = Font(name="Microsoft YaHei", bold=True, size=11)
    sub_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    normal_font = Font(name="Microsoft YaHei", size=11)
    note_font = Font(name="Microsoft YaHei", size=10, color="666666")
    note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = Workbook()

    # ════════════════════════════════════════════
    # Sheet 1: 导入模板
    # ════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "导入模板"

    # Row 1: English headers
    headers_en = ["filename", "material_brand", "material_type", "color", "quantity", "printer", "nozzle", "layer_height", "wall_count", "infill"]
    for col, val in enumerate(headers_en, 1):
        cell = ws1.cell(row=1, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Row 2: Chinese sub-headers
    headers_cn = ["文件名", "材料品牌", "材料", "颜色", "数量", "打印机", "喷嘴直径", "层高(mm)", "墙层数", "填充密度(%)"]
    for col, val in enumerate(headers_cn, 1):
        cell = ws1.cell(row=2, column=col, value=val)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Build brand display strings from user's brands
    brand_examples = user_brands[:3] if len(user_brands) >= 3 else user_brands
    # Pad to at least 3 for the example rows
    while len(brand_examples) < 3:
        brand_examples.append("Generic")

    # Row 3-5: Example data
    examples = [
        ["model1.stl", brand_examples[0], "PLA", "白色", 1, "Bambu Lab A1", 0.4, 0.2, 3, 20],
        ["model2.stl", "", "", "", "", "", 0.16, 4, 15],
        ["model3.stl", brand_examples[2] if len(brand_examples) > 2 else "Generic", "PETG", "黑色", 2, "Creality K1 Max", 0.6, 0.28, 2, 10],
    ]
    for r, row_data in enumerate(examples, 3):
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=col, value=val if val != "" else None)
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = thin_border

    # Note row explaining empty = default
    note_row = 6
    note_text = "💡 提示：空白单元格 = 使用系统默认值，填写 = 覆盖默认值。第一行（英文列名）必须保留。"
    ws1.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
    note_cell = ws1.cell(row=note_row, column=1, value=note_text)
    note_cell.font = note_font
    note_cell.fill = note_fill
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for c in range(1, 11):
        ws1.cell(row=note_row, column=c).border = thin_border

    # Auto-width for sheet 1
    col_widths_s1 = [16, 14, 12, 10, 10, 20, 14, 16, 12, 16]
    for i, w in enumerate(col_widths_s1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════
    # Sheet 2: 参数说明
    # ════════════════════════════════════════════
    ws2 = wb.create_sheet("参数说明")

    # Headers
    param_headers = ["参数", "英文列名", "说明", "默认值", "可选值"]
    for col, val in enumerate(param_headers, 1):
        cell = ws2.cell(row=1, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Build brand list string for parameter description
    brands_display = ", ".join(user_brands)

    # Parameter rows
    params = [
        ["文件名", "filename", "必填，与压缩包内模型文件名（不含扩展名）匹配", "—", "—"],
        ["材料品牌", "material_brand", "可选，耗材品牌名称", "Generic", brands_display],
        ["材料", "material_type", "可选，材料类型（如 PETG, PLA, ABS）", "—", "PETG, PLA, PLA+, ABS, ASA, TPU, PA, PC"],
        ["颜色", "color", "可选，模型颜色", "使用表单默认", "白色, 黑色, 红色, 蓝色, 绿色 等"],
        ["数量", "quantity", "可选，正整数", "使用表单默认", "1, 2, 3, ..."],
        ["打印机", "printer", "可选，打印机型号", "使用系统默认", "Bambu Lab A1, Creality K1, Prusa MK4 等"],
        ["喷嘴直径", "nozzle", "可选，单位 mm", "0.4", "0.2, 0.4, 0.6, 0.8"],
        ["层高", "layer_height", "可选，单位 mm", "0.2", "0.08, 0.10, 0.12, 0.16, 0.20, 0.28, 0.32"],
        ["墙层数", "wall_count", "可选，外壁层数", "3", "2, 3, 4, 5, 6, 8"],
        ["填充密度", "infill", "可选，百分比", "20", "5, 10, 15, 20, 25, 30, 40, 50, 60, 80, 100"],
    ]
    for r, row_data in enumerate(params, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=col, value=val)
            cell.font = normal_font
            cell.alignment = left_align if col >= 3 else center_align
            cell.border = thin_border

    # Auto-width for sheet 2
    col_widths_s2 = [12, 16, 40, 16, 44]
    for i, w in enumerate(col_widths_s2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Save to buffer and return ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=zip_import_template.xlsx"},
    )
